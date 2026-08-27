"""
Reads every data row from 'All SKU and Details 1.xlsx' (Sheet1, headers on
row 3, data from row 4 to row 2594) and writes plating_stock_master_data.csv
with only the columns that map to real ModelMaster / JigLoadingMaster
fields. No values are invented - a blank/missing cell in the source is
written through as an empty string, never guessed.

Usage:
    python extract_plating_stock_master.py
"""
import csv
import openpyxl

SRC_PATH = "All SKU and Details 1.xlsx"
OUT_PATH = "plating_stock_master_data.csv"

wb = openpyxl.load_workbook(SRC_PATH, read_only=True, data_only=True)
ws = wb["Sheet1"]

rows = list(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True))

out_rows = []
skipped_blank = 0
for r in rows:
    plating_stk_no = r[0]
    if plating_stk_no is None or str(plating_stk_no).strip() == "":
        skipped_blank += 1
        continue

    basic_stock_no = r[3]
    basic_stk_version = r[5]
    plating_stk_no = str(plating_stk_no).strip()
    basic_stock_no = "" if basic_stock_no is None else str(basic_stock_no).strip()
    basic_stk_version = "" if basic_stk_version is None else str(basic_stk_version).strip()

    version = ""
    if basic_stk_version and basic_stock_no and basic_stk_version.startswith(basic_stock_no):
        version = basic_stk_version[len(basic_stock_no):]
    elif basic_stk_version:
        version = basic_stk_version

    def s(v):
        return "" if v is None else str(v).strip()

    out_rows.append({
        "plating_stk_no": plating_stk_no,
        "model_no": basic_stock_no,
        "version": version,
        "brand": s(r[6]),
        "gender": s(r[7]),
        "pol_finish": s(r[9]),
        "jig_qty": s(r[11]),
        "traycate": s(r[12]),
        "ip_wiping": s(r[13]),
        "bath_type": s(r[14]),
    })

with open(OUT_PATH, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=[
        "plating_stk_no", "model_no", "version", "brand", "gender",
        "pol_finish", "jig_qty", "traycate", "ip_wiping", "bath_type",
    ])
    writer.writeheader()
    writer.writerows(out_rows)

print(f"Total data rows in sheet (row 4 to {ws.max_row}): {len(rows)}")
print(f"Rows with blank Plating Stock No (skipped): {skipped_blank}")
print(f"Rows written to {OUT_PATH}: {len(out_rows)}")
