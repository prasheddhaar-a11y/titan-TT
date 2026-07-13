import logging
logger = logging.getLogger(__name__)
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils.safestring import mark_safe
from django.db.models import OuterRef, Subquery, Q
from django.db import transaction, IntegrityError
from django.utils.timezone import now
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from datetime import timedelta, datetime
import datetime as _dt
import json
import re
import math
import openpyxl
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.renderers import TemplateHTMLRenderer
from rest_framework.parsers import MultiPartParser, JSONParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import api_view, permission_classes
from .models import *
from adminportal.views import *
from modelmasterapp.models import RowAccessLock, DraftTrayId
from watchcase_tracker.perf_logger import time_stage

# ── Pre-compiled regex patterns (compiled once at import, reused every row) ────
_RE_STOCK = re.compile(r'^(\d+)([A-Z])([A-Z][A-Z]02)$')
_RE_SUFFIX = re.compile(r'^[A-Z][A-Z]02$')



# API to lock a row when accessed - Day Planning Pick Table

@csrf_exempt
def lock_row_api(request):
    if request.method == "POST":
        batch_id = request.POST.get("batch_id")
        lot_id = request.POST.get("lot_id")
        user = request.user
        action = request.POST.get("action", "unknown")
        if not batch_id or not user.is_authenticated:
            return JsonResponse({"success": False, "error": "Missing data"}, status=400)
        print(f"[ROW LOCK] {now().strftime('%d/%b/%Y %H:%M:%S')}]")
        print(f"BATCH-ID - {batch_id}")
        print(f"Lot ID - {lot_id if lot_id else 'not exists'}")
        print(f"User name - {user.username if user else 'Anonymous'}")
        print(f"Action performing - {action}")
        print(f"Start Accessing status - active")
        print(f"End accessing status - not ended yet")
        print(f'"POST /dayplanning/row_lock/ batch_id={batch_id} lot_id={lot_id if lot_id else "null"} Action={action} Status=actively accessing"')

        obj, created = RowAccessLock.objects.get_or_create(
            batch_id=batch_id, lot_id=lot_id,
            defaults={"accessed_by": user}
        )

        if not created and obj.accessed_by != user:
            return JsonResponse({
                "success": False,
                "locked": True,
                "by": obj.accessed_by.username
            })

        # ✅ owner always gets "by" = himself
        return JsonResponse({
            "success": True,
            "locked": True,
            "by": user.username
        })


# When the owner closes the modal or finishes editing, release the lock
def unlock_row_api(request):
    if request.method == "POST":
        batch_id = request.POST.get("batch_id")
        lot_id = request.POST.get("lot_id")
        user = request.user
        if not batch_id or not user.is_authenticated:
            return JsonResponse({"success": False, "error": "Missing data"}, status=400)
        obj = RowAccessLock.objects.filter(batch_id=batch_id, lot_id=lot_id, accessed_by=user).first()
        if obj:
            obj.delete()
            return JsonResponse({"success": True, "unlocked": True})
        return JsonResponse({"success": False, "error": "No lock found"}, status=404)


@csrf_exempt
def check_row_lock_api(request):
    batch_id = request.GET.get("batch_id")
    lot_id = request.GET.get("lot_id")
    user = request.user if request.user.is_authenticated else None
    if not batch_id and not lot_id:
        print(f"[ROW LOCK] {now().strftime('%d/%b/%Y %H:%M:%S')}]")
        print(f"BATCH-ID - None")
        print(f"Lot ID - not exists")
        print(f"User name - {user.username if user else 'Anonymous'}")
        print(f"Action performing - check")
        print(f"Start Accessing status - not started")
        return JsonResponse({"success": False, "error": "Missing data"}, status=400)
    filters = {}
    if batch_id:
        filters['batch_id'] = batch_id
    if lot_id:
        filters['lot_id'] = lot_id
    obj = RowAccessLock.objects.filter(**filters).first()
    print(f"[ROW LOCK] {now().strftime('%d/%b/%Y %H:%M:%S')}]")
    print(f"BATCH-ID - {batch_id}")
    print(f"Lot ID - {lot_id if lot_id else 'not exists'}")
    print(f"User name - {user.username if user else 'Anonymous'}")
    print(f"Action performing - check")
    print(f"Start Accessing status - active")

    if obj:
        return JsonResponse({
            "success": True,
            "locked": True,
            "by": obj.accessed_by.username if obj.accessed_by else "Unknown"
        })
    return JsonResponse({"success": True, "locked": False})


@method_decorator(csrf_exempt, name='dispatch')
class DPQuickHelpAPIView(APIView):
    """
    Fetch Day Planning Quick Help Do's and Don'ts from database
    Real-time content - admins can add/edit/delete from Django admin panel
    """
    
    def get(self, request):
        try:
            # Fetch active guidelines, grouped by category
            dos = DPQuickHelp.objects.filter(
                category='do', 
                is_active=True
            ).order_by('order', 'created_at').values('title', 'description', 'icon_code')
            
            donts = DPQuickHelp.objects.filter(
                category='dont', 
                is_active=True
            ).order_by('order', 'created_at').values('title', 'description', 'icon_code')
            
            return JsonResponse({
                'success': True,
                'dos': list(dos),
                'donts': list(donts)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': 'Unable to process the request. Please verify the submitted data and try again.'
            }, status=500)


class DPBulkUploadView(APIView):
    """
    Enhanced Day Planning Bulk Upload View
    
    Handles two source formats:
    1. Vendor_Location format (e.g., "Titan_CPSE") - splits into vendor and location
    2. Location only format (e.g., "CPSE") - treats entire value as location name
    
    Supports ambiguous plating color resolution using additional identifiers
    after "/" in plating stock numbers (e.g., "2648QAA02/BRN").
    
    Enhanced with comprehensive column validation and detailed error reporting.
    """
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'Day_Planning/DP_BulkUpload.html'
    parser_classes = [MultiPartParser, JSONParser]
    permission_classes = [IsAuthenticated] 

    def validate_excel_columns(self, sheet):
        """
        Validate Excel file columns to ensure they match expected format
        Returns: (is_valid, error_message, header_row)
        """
        try:
            # Expected column names in exact order
            expected_columns = [
                'S.No',
                'Plating Stk No', 
                'Polishing Stk No',
                'Plating Colour',
                'Category',
                'Input Qty',
                'Source'
            ]
            
            # Get the first row (header row)
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            
            if not header_row:
                return False, "❌ Excel file appears to be empty or corrupted.", None
            
            # Convert header values to strings and strip whitespace
            actual_columns = [str(cell).strip() if cell is not None else '' for cell in header_row]
            
            # Remove empty columns from the end
            while actual_columns and actual_columns[-1] == '':
                actual_columns.pop()
            
            # Check if we have the minimum required number of columns
            if len(actual_columns) < len(expected_columns):
                missing_count = len(expected_columns) - len(actual_columns)
                missing_columns = expected_columns[len(actual_columns):]
                return False, f"❌ Missing {missing_count} column(s): {', '.join(missing_columns)}. Expected {len(expected_columns)} columns but found {len(actual_columns)}.", actual_columns
            
            # Check for exact column name matches
            mismatched_columns = []
            for i, (expected, actual) in enumerate(zip(expected_columns, actual_columns)):
                if expected != actual:
                    mismatched_columns.append(f"Column {i+1}: Expected '{expected}' but found '{actual}'")
            
            if mismatched_columns:
                error_msg = "❌ Column Name Mismatch:\n" + "\n".join(mismatched_columns)
                error_msg += f"\n\nExpected format: {' | '.join(expected_columns)}"
                error_msg += f"\nActual format: {' | '.join(actual_columns[:len(expected_columns)])}"
                return False, error_msg, actual_columns
            
            # Check for extra columns (warn but don't fail)
            if len(actual_columns) > len(expected_columns):
                extra_columns = actual_columns[len(expected_columns):]
                print(f"⚠️ Warning: Found {len(extra_columns)} extra column(s) that will be ignored: {', '.join(extra_columns)}")
            
            return True, "✅ Column validation passed", actual_columns
            
        except Exception as e:
            return False, f"❌ Error reading Excel file headers: {str(e)}", None




    def validate_codes(self, plating_stock_no, polishing_stock_no, model_masters=None, polish_types=None, versions=None):
        """
        Enhanced validation to handle ambiguous plating codes.
        
        When a plating code (like 'Q') maps to multiple colors in the database,
        this method looks for additional identifier after "/" in plating_stock_no.
        
        Example: "2648QAA02/BRN" 
        - Q is the plating code but ambiguous
        - /BRN is the specific identifier to match plating_color_internal='BRN'
        
        IMPORTANT: Stock number validation rules:
        - Plating: 1805SAB02 (any UPPERCASE color code, must end with 02)
        - Polishing: 1805XAB02 (must have UPPERCASE X as color code, must end with 02)
        - Model numbers must match: 1805 = 1805
        - ALL LETTERS MUST BE UPPERCASE: SAB02 ✅, sab02 ❌, SAb02 ❌
        - Suffixes must match except color: SAB02 vs XAB02 (only S vs X difference)
        
        Returns: (model_stock, codes_tuple, error_msg)
        """
        try:
            # STEP 0: Validate that both stock numbers end with "02"
            plating_base = plating_stock_no.split("/")[0] if "/" in plating_stock_no else plating_stock_no
            polishing_base = polishing_stock_no.split("/")[0] if "/" in polishing_stock_no else polishing_stock_no
            
            if not plating_base.endswith("02"):
                return None, None, f"❌ Invalid Plating Stk No format: '{plating_stock_no}' must end with '02'. Example: 1805SAB02"
            
            if not polishing_base.endswith("02"):
                return None, None, f"❌ Invalid Polishing Stk No format: '{polishing_stock_no}' must end with '02'. Example: 1805XAB02"

            # STEP 0.1: CRITICAL - Validate that ALL letters are UPPERCASE
            # Use filter+isupper — faster than re.findall for this simple check
            plating_alpha = [c for c in plating_base if c.isalpha()]
            if any(c.islower() for c in plating_alpha):
                lowercase_found = [c for c in plating_alpha if c.islower()]
                return None, None, f"❌ Invalid Plating Stk No: '{plating_stock_no}' contains lowercase letters '{', '.join(lowercase_found)}'. ALL letters must be UPPERCASE. Correct format: {plating_base.upper()}"

            polishing_alpha = [c for c in polishing_base if c.isalpha()]
            if any(c.islower() for c in polishing_alpha):
                lowercase_found = [c for c in polishing_alpha if c.islower()]
                return None, None, f"❌ Invalid Polishing Stk No: '{polishing_stock_no}' contains lowercase letters '{', '.join(lowercase_found)}'. ALL letters must be UPPERCASE. Correct format: {polishing_base.upper()}"

            # STEP 1: Validate that plating and polishing stock numbers match pattern
            # Use module-level pre-compiled patterns (avoids recompile every call)
            plating_match = _RE_STOCK.match(plating_base)
            polishing_match = _RE_STOCK.match(polishing_base)
            
            if not plating_match:
                return None, None, f"❌ Invalid Plating Stk No format: '{plating_stock_no}'. Expected format: ModelNumber + UPPERCASE_ColorCode + UPPERCASE_Letters + 02 (e.g., 1805SAB02). Current: {plating_base}"
            
            if not polishing_match:
                return None, None, f"❌ Invalid Polishing Stk No format: '{polishing_stock_no}'. Expected format: ModelNumber + X + UPPERCASE_Letters + 02 (e.g., 1805XAB02). Current: {polishing_base}"
            
            plating_model, plating_color_code, plating_suffix = plating_match.groups()
            polishing_model, polishing_color_code, polishing_suffix = polishing_match.groups()
            
            # STEP 1.1: Validate that polishing stock number has "X" as color code (UPPERCASE)
            if polishing_color_code != 'X':
                return None, None, f"❌ Invalid Polishing Stk No: '{polishing_stock_no}'. Polishing stock number must have UPPERCASE 'X' as color code. Expected: {polishing_model}X{polishing_suffix} (not '{polishing_color_code}')"
            
            # STEP 1.2: Validate that model numbers match
            if plating_model != polishing_model:
                return None, None, f"❌ Stock number mismatch: Plating model '{plating_model}' does not match Polishing model '{polishing_model}'. Expected format: {plating_model}[COLOR]AB02 for plating and {plating_model}XAB02 for polishing."
            
            # STEP 1.3: Create expected polishing suffix by replacing color code with X
            expected_polishing_suffix = plating_suffix
            expected_polishing_stock = f"{plating_model}X{expected_polishing_suffix}"
            
            # Validate that polishing suffix matches plating suffix (except for color code)
            if polishing_suffix != expected_polishing_suffix:
                return None, None, f"❌ Stock number mismatch: Expected polishing stock number '{expected_polishing_stock}' but got '{polishing_stock_no}'. Only the color code should be 'X'."
            
            # STEP 1.4: Additional validation: Ensure suffix follows UPPERCASE AB02 pattern
            if not _RE_SUFFIX.match(plating_suffix):
                return None, None, f"❌ Invalid suffix pattern in Plating Stk No: '{plating_suffix}'. Expected pattern: [UPPERCASE][UPPERCASE]02 (e.g., AB02, not ab02 or Ab02)"

            if not _RE_SUFFIX.match(polishing_suffix):
                return None, None, f"❌ Invalid suffix pattern in Polishing Stk No: '{polishing_suffix}'. Expected pattern: [UPPERCASE][UPPERCASE]02 (e.g., AB02, not ab02 or Ab02)"

            # STEP 2: Extract model number and validate in ModelMaster
            model_no = plating_model
            if model_masters is not None:
                model_stock = model_masters.get(model_no)
            else:
                model_stock = ModelMaster.objects.filter(model_no=model_no).first()
            if not model_stock:
                return None, None, f"❌ Plating Stk No '{plating_stock_no}' - Model number '{model_no}' not available in Master Data."

            # STEP 3: Determine plating color internal code
            if "/" in plating_stock_no:
                plating_color_internal = plating_stock_no.split("/")[1]
            else:
                plating_color_internal = plating_color_code

            # STEP 4: Extract polish and version codes from polishing stock number
            letters = ''.join(c for c in polishing_stock_no if c.isupper())
            if len(letters) < 3:
                return None, None, f"❌ Invalid polishing stock format. Found less than 3 UPPERCASE letters in: {polishing_stock_no}. Expected format: ModelNumber + X + [UPPERCASE][UPPERCASE]02"

            # Since we know polishing has X as color code, extract polish and version codes
            polish_code = letters[1]  # Second letter for polish (after X)
            version_code = letters[2]  # Third letter for version

            # STEP 4.1: Validate polish code against PolishFinishType master data
            if polish_types is not None:
                polish_obj = polish_types.get(polish_code)
            else:
                polish_obj = PolishFinishType.objects.filter(polish_internal=polish_code).first()
            if not polish_obj:
                available_polish = list(PolishFinishType.objects.values_list('polish_internal', flat=True))
                polish_suggestion = f" Available polish codes: {', '.join(available_polish)}" if available_polish else ""
                return None, None, f"❌ Invalid Polish Code '{polish_code}' in Polishing Stk No '{polishing_stock_no}'.{polish_suggestion}"

            # STEP 4.2: Validate version code against Version master data
            if versions is not None:
                version_obj = versions.get(version_code)
            else:
                version_obj = Version.objects.filter(
                    Q(version_internal=version_code) | 
                    Q(version_name=version_code)
                ).first()
            
            if not version_obj:
                available_versions_internal = list(Version.objects.exclude(version_internal__isnull=True).exclude(version_internal='').values_list('version_internal', flat=True))
                available_versions_name = list(Version.objects.values_list('version_name', flat=True))
                all_available_versions = list(set(available_versions_internal + available_versions_name))
                version_suggestion = f" Available version codes: {', '.join(sorted(all_available_versions)[:10])}" if all_available_versions else ""
                return None, None, f"❌ Invalid Version Code '{version_code}' in Polishing Stk No '{polishing_stock_no}'.{version_suggestion}"

            # Return as tuple of 3 values: model_stock, codes_tuple, error_msg
            codes_tuple = (plating_color_internal, polish_code, version_code)
            return model_stock, codes_tuple, None

        except Exception as e:
            return None, None, f"❌ Error validating codes: {str(e)}"

    def get(self, request, format=None):
        master_data = ModelMasterCreation.objects.all()
        return Response({'master_data': master_data})

    def post(self, request, format=None):
        # Check if this is a preview request
        if 'preview' in request.path:
            return self.handle_file_preview(request)
        
        # Check if request contains JSON data (from datatable)
        if request.content_type == 'application/json':
            return self.handle_datatable_submission(request)
        
        # Handle file upload (existing functionality)
        return self.handle_file_upload(request)

    def handle_file_preview(self, request):
        """Handle file preview for datatable with column validation"""
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({
                'success': False,
                'error': '❌ No file uploaded.'
            }, status=400)

        if not uploaded_file.name.endswith(('.xls', '.xlsx')):
            return JsonResponse({
                'success': False,
                'error': f"❌ Only Excel files are allowed. '{uploaded_file.name}' is not valid."
            }, status=400)

        try:
            wb = None
            wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
            sheet = wb.active
            if not sheet:
                return JsonResponse({
                    'success': False,
                    'error': '❌ Could not read the Excel sheet.'
                }, status=400)

            # ========== VALIDATE COLUMNS FIRST ==========
            is_valid, error_message, actual_columns = self.validate_excel_columns(sheet)
            if not is_valid:
                return JsonResponse({
                    'success': False,
                    'error': error_message
                }, status=400)
            
            print("✅ Excel column validation passed for preview")
            # ========== END COLUMN VALIDATION ==========

            # Process rows for preview
            preview_data = []
            row_errors = []
            
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                # Check for minimum required columns
                if len(row) < 7:
                    row_errors.append(f"Row {idx}: ❌ Missing Data — Expected 7 columns but found {len(row)}")
                    continue

                # Extract data
                s_no = str(row[0]).strip() if row[0] else ''
                plating_stock_no = str(row[1]).strip() if row[1] else ''
                polishing_stock_no = str(row[2]).strip() if row[2] else ''
                plating_colour = str(row[3]).strip() if row[3] else ''
                category = str(row[4]).strip() if row[4] else ''
                input_qty = row[5]
                source = str(row[6]).strip() if row[6] else ''

                # Basic validation for preview
                row_data = {
                    'S.No': s_no or str(len(preview_data) + 1),
                    'Plating Stk No': plating_stock_no,
                    'Polishing Stk No': polishing_stock_no,
                    'Plating Colour': plating_colour,
                    'Category': category,
                    'Input Qty': input_qty,
                    'Source': source
                }
                
                preview_data.append(row_data)
                


            if not preview_data and row_errors:
                return JsonResponse({
                    'success': False,
                    'error': f"❌ No valid data found.\n\n" + "\n".join(row_errors[:5])
                }, status=400)

            response_data = {
                'success': True,
                'data': preview_data
            }
            
            if row_errors:
                response_data['warnings'] = row_errors[:10]  # Show first 10 warnings
                
            return JsonResponse(response_data)

        except Exception as e:
            logger.error(f"❌ Error processing file preview: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': 'Unable to process the request. Please verify the submitted data and try again.'
            }, status=500)
        finally:
            if wb:
                wb.close()

    def handle_datatable_submission(self, request):
        """Handle submission from HTML datatable — optimized for bulk uploads (1000+ rows)."""
        try:
            data = request.data
            rows = data.get('rows', [])

            if not rows:
                return JsonResponse({
                    'success': False,
                    'error': '❌ No data provided for processing.'
                }, status=400)

            # ── Pre-fetch ALL master data once ─────────────────────────────────────
            # select_related('tray_type') prevents N+1 lazy FK queries per row
            model_masters = {obj.model_no: obj for obj in ModelMaster.objects.select_related('tray_type').all()}
            polish_types = {obj.polish_internal: obj for obj in PolishFinishType.objects.all()}
            versions = {}
            for obj in Version.objects.all():
                if obj.version_internal:
                    versions[obj.version_internal] = obj
                versions[obj.version_name] = obj
            plating_colors_internal = {}
            for obj in Plating_Color.objects.order_by('id'):
                if obj.plating_color_internal not in plating_colors_internal:
                    plating_colors_internal[obj.plating_color_internal] = obj
            plating_colors_name = {obj.plating_color: obj for obj in Plating_Color.objects.all()}
            categories = {obj.category_name: obj for obj in Category.objects.all()}
            vendors = {obj.vendor_name: obj for obj in Vendor.objects.all()}
            locations = {obj.location_name: obj for obj in Location.objects.all()}

            # Pre-build suggestion strings from cached dicts (avoids per-error DB hits)
            _color_hint = ", ".join(list(plating_colors_name.keys())[:5])
            _polish_hint = ", ".join(list(polish_types.keys())[:5])
            _version_hint = ", ".join(sorted(list(versions.keys()))[:10])
            _cat_hint = ", ".join(list(categories.keys())[:5])
            _vendor_hint = ", ".join(list(vendors.keys())[:5])
            _loc_hint = ", ".join(list(locations.keys())[:5])
            # ───────────────────────────────────────────────────────────────────────

            success_count = 0
            failure_count = 0
            failed_rows = []
            objects_to_create = []

            # Single timestamp prefix for all batch IDs in this upload
            import datetime as dt
            upload_ts = dt.datetime.now().strftime('%Y%m%d%H%M%S%f')

            # Per-upload cache: avoid re-running validate_codes for duplicate stock pairs
            _codes_cache = {}

            for idx, row_data in enumerate(rows, start=1):
                try:
                    plating_stock_no = str(row_data.get('Plating Stk No', '')).strip()
                    polishing_stock_no = str(row_data.get('Polishing Stk No', '')).strip()
                    plating_colour = str(row_data.get('Plating Colour', '')).strip()
                    category = str(row_data.get('Category', '')).strip()
                    input_qty = row_data.get('Input Qty')
                    source = str(row_data.get('Source', '')).strip()

                    # Validate required fields
                    empty_fields = []
                    if not plating_stock_no:   empty_fields.append("Plating Stk No")
                    if not polishing_stock_no: empty_fields.append("Polishing Stk No")
                    if not plating_colour:     empty_fields.append("Plating Colour")
                    if not category:           empty_fields.append("Category")
                    if input_qty in [None, '', 0]: empty_fields.append("Input Qty")
                    if not source:             empty_fields.append("Source")

                    if empty_fields:
                        failed_rows.append(f"Row {idx}: ❌ {', '.join(empty_fields)} should not be empty.")
                        failure_count += 1
                        continue

                    try:
                        input_qty = int(input_qty)
                    except (ValueError, TypeError):
                        failed_rows.append(f"Row {idx}: ❌ Invalid quantity value: {input_qty}")
                        failure_count += 1
                        continue

                    cache_key = (plating_stock_no, polishing_stock_no)
                    if cache_key in _codes_cache:
                        model_stock, codes, error_msg = _codes_cache[cache_key]
                    else:
                        model_stock, codes, error_msg = self.validate_codes(
                            plating_stock_no, polishing_stock_no, model_masters, polish_types, versions
                        )
                        _codes_cache[cache_key] = (model_stock, codes, error_msg)
                    if error_msg:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: {error_msg}")
                        continue
                    if not codes:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Could not extract codes from stock numbers.")
                        continue

                    plating_color_internal, polish_code, version_code = codes

                    # 1. Plating internal code
                    plating_obj_code = plating_colors_internal.get(plating_color_internal)
                    if not plating_obj_code:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Plating color internal code '{plating_color_internal}' not available in Master Data.")
                        continue

                    # 2. Plating colour from Excel
                    plating_color_obj = plating_colors_name.get(plating_colour)
                    if not plating_color_obj:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Plating Colour '{plating_colour}' not available in Master Data. Available: {_color_hint}")
                        continue

                    # 3. Cross-validate internal code vs Excel colour
                    if plating_obj_code.pk != plating_color_obj.pk:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Plating color mismatch: Stock code '{plating_stock_no}' resolves to '{plating_obj_code.plating_color}' but Excel shows '{plating_colour}'.")
                        continue

                    # 4. Polish code
                    polish_obj = polish_types.get(polish_code)
                    if not polish_obj:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Polish code '{polish_code}' not available in Master Data. Available: {_polish_hint}")
                        continue

                    # 5. Version code
                    version_obj = versions.get(version_code)
                    if not version_obj:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Version code '{version_code}' not available in Master Data. Available: {_version_hint}")
                        continue

                    # 6. Category
                    category_obj = categories.get(category)
                    if not category_obj:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Category '{category}' not available in Master Data. Available: {_cat_hint}")
                        continue

                    # 7. Source (Vendor_Location or Location-only)
                    vendor_obj = None
                    location_obj = None
                    if "_" in source:
                        vendor_name, loc_name = source.split("_", 1)
                        vendor_obj = vendors.get(vendor_name)
                        if not vendor_obj:
                            failure_count += 1
                            failed_rows.append(f"Row {idx}: ❌ Vendor '{vendor_name}' not available in Master Data. Available: {_vendor_hint}")
                            continue
                        location_obj = locations.get(loc_name)
                        if not location_obj:
                            failure_count += 1
                            failed_rows.append(f"Row {idx}: ❌ Location '{loc_name}' not available in Master Data. Available: {_loc_hint}")
                            continue
                    else:
                        location_obj = locations.get(source)
                        if not location_obj:
                            failure_count += 1
                            failed_rows.append(f"Row {idx}: ❌ Location '{source}' not available in Master Data. Available: {_loc_hint}")
                            continue

                    # Collect validated object — DO NOT save yet
                    objects_to_create.append(ModelMasterCreation(
                        batch_id=f"BATCH-{upload_ts}-{idx}",
                        model_stock_no=model_stock,
                        plating_color=plating_obj_code.plating_color,
                        vendor_internal=vendor_obj.vendor_internal if vendor_obj else None,
                        location=location_obj,
                        tray_capacity=model_stock.tray_capacity if model_stock else None,
                        tray_type=model_stock.tray_type.tray_type if model_stock and model_stock.tray_type else None,
                        ep_bath_type=model_stock.ep_bath_type if model_stock else None,
                        total_batch_quantity=input_qty,
                        version=version_obj,
                        polish_finish=polish_obj,
                        category=category_obj,
                        plating_stk_no=plating_stock_no,
                        polishing_stk_no=polishing_stock_no,
                    ))
                    success_count += 1

                except Exception as e:
                    failure_count += 1
                    failed_rows.append("Row processing failed. Please verify the row data and try again.")

            # ── Single bulk INSERT inside one transaction ───────────────────────────
            if objects_to_create:
                last_sequence = ModelMasterCreation.objects.aggregate(
                    max_seq=models.Max('sequence_number')
                )['max_seq'] or 0
                for i, obj in enumerate(objects_to_create, start=1):
                    obj.sequence_number = last_sequence + i
                with transaction.atomic():
                    ModelMasterCreation.objects.bulk_create(objects_to_create)
                print(f"✅ Bulk inserted {len(objects_to_create)} records")
            # ───────────────────────────────────────────────────────────────────────

            if success_count > 0 and failure_count == 0:
                return JsonResponse({
                    'success': True,
                    'message': f"✅ {success_count} row(s) processed successfully."
                })
            elif success_count > 0 and failure_count > 0:
                return JsonResponse({
                    'success': True,
                    'message': f"⚠️ Partial Success: {success_count} succeeded, {failure_count} failed.",
                    'failed_rows': failed_rows
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': f"❌ All {failure_count} row(s) failed to process.",
                    'failed_rows': failed_rows
                }, status=400)

        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': '❌ Invalid JSON data provided.'
            }, status=400)
        except Exception as e:
            logger.error(f"❌ Error in datatable submission: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': 'Unable to process the request. Please verify the submitted data and try again.'
            }, status=500)

    def handle_file_upload(self, request):
        """Handle file upload with enhanced column validation"""
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            messages.error(request, "❌ No file uploaded.")
            return Response({'master_data': ModelMasterCreation.objects.none()}, status=status.HTTP_400_BAD_REQUEST)

        if not uploaded_file.name.endswith(('.xls', '.xlsx')):
            messages.error(request, f"❌ Only Excel files are allowed. '{uploaded_file.name}' is not valid.")
            return Response({'master_data': ModelMasterCreation.objects.none()}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = None
            wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
            sheet = wb.active
            if not sheet:
                messages.error(request, "❌ Could not read the Excel sheet.")
                return Response({'master_data': ModelMasterCreation.objects.none()}, status=status.HTTP_400_BAD_REQUEST)

            # ========== VALIDATE COLUMNS BEFORE PROCESSING ==========
            is_valid, error_message, actual_columns = self.validate_excel_columns(sheet)
            if not is_valid:
                messages.error(request, mark_safe(error_message.replace('\n', '<br>')))
                return Response({'master_data': ModelMasterCreation.objects.none()}, status=status.HTTP_400_BAD_REQUEST)
            
            print("✅ Excel column validation passed")
            # ========== END COLUMN VALIDATION ==========

            success_count = 0
            failure_count = 0
            failed_rows = []
            objects_to_create = []

            # Pre-fetch all master data for performance optimization
            # select_related('tray_type') prevents N+1 lazy FK queries per row
            model_masters = {obj.model_no: obj for obj in ModelMaster.objects.select_related('tray_type').all()}
            polish_types = {obj.polish_internal: obj for obj in PolishFinishType.objects.all()}
            versions = {}
            for obj in Version.objects.all():
                if obj.version_internal:
                    versions[obj.version_internal] = obj
                versions[obj.version_name] = obj
            plating_colors_internal = {obj.plating_color_internal: obj for obj in Plating_Color.objects.all()}
            plating_colors_name = {obj.plating_color: obj for obj in Plating_Color.objects.all()}
            categories = {obj.category_name: obj for obj in Category.objects.all()}
            vendors = {obj.vendor_name: obj for obj in Vendor.objects.all()}
            locations = {obj.location_name: obj for obj in Location.objects.all()}

            # Pre-build suggestion strings (avoids per-error DB queries inside the loop)
            _color_hint = ", ".join(list(plating_colors_name.keys())[:5])
            _polish_hint = ", ".join(list(polish_types.keys())[:5])
            _version_hint = ", ".join(sorted(list(versions.keys()))[:10])
            _cat_hint = ", ".join(list(categories.keys())[:5])
            _vendor_hint = ", ".join(list(vendors.keys())[:5])
            _loc_hint = ", ".join(list(locations.keys())[:5])

            # Single timestamp prefix for all batch IDs in this upload
            import datetime as dt
            upload_ts = dt.datetime.now().strftime('%Y%m%d%H%M%S%f')

            # Per-upload cache: avoid re-running validate_codes for duplicate stock pairs
            _codes_cache = {}

            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                # ========== ENHANCED: CHECK FOR MINIMUM REQUIRED COLUMNS ==========
                if len(row) < 7:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Missing Data — Expected 7 columns but found {len(row)}. Please ensure all required columns are filled.")
                    continue
                # ========== END ENHANCED COLUMN CHECK ==========

                # Extract data with new field structure
                s_no = str(row[0]).strip() if row[0] else ''
                plating_stock_no = str(row[1]).strip() if row[1] else ''
                polishing_stock_no = str(row[2]).strip() if row[2] else ''
                plating_colour = str(row[3]).strip() if row[3] else ''
                category = str(row[4]).strip() if row[4] else ''
                input_qty = row[5]
                source = str(row[6]).strip() if row[6] else ''

                # ========== ENHANCED: DETAILED EMPTY FIELD VALIDATION ==========
                empty_fields = []
                if not plating_stock_no:
                    empty_fields.append("Column B (Plating Stk No)")
                if not polishing_stock_no:
                    empty_fields.append("Column C (Polishing Stk No)")
                if not plating_colour:
                    empty_fields.append("Column D (Plating Colour)")
                if not category:
                    empty_fields.append("Column E (Category)")
                if input_qty in [None, '', 0]:
                    empty_fields.append("Column F (Input Qty)")
                if not source:
                    empty_fields.append("Column G (Source)")

                if empty_fields:
                    field_list = ", ".join(empty_fields)
                    failed_rows.append(f"Row {idx}: ❌ Empty Required Fields — {field_list} cannot be empty.")
                    failure_count += 1
                    continue
                # ========== END ENHANCED EMPTY FIELD VALIDATION ==========

                # Convert quantity to integer
                try:
                    input_qty = int(input_qty)
                except (ValueError, TypeError):
                    failed_rows.append(f"Row {idx}: ❌ Invalid Data — Column F (Input Qty): '{input_qty}' is not a valid number.")
                    failure_count += 1
                    continue

                # Validate codes using new stock numbers
                cache_key = (plating_stock_no, polishing_stock_no)
                if cache_key in _codes_cache:
                    model_stock, codes, error_msg = _codes_cache[cache_key]
                else:
                    model_stock, codes, error_msg = self.validate_codes(
                        plating_stock_no, polishing_stock_no, model_masters, polish_types, versions
                    )
                    _codes_cache[cache_key] = (model_stock, codes, error_msg)
                if error_msg:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: {error_msg}")
                    continue

                if not codes:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Could not extract codes from stock numbers.")
                    continue

                plating_color_internal, polish_code, version_code = codes

                # ========== APPLY SAME VALIDATION AS DATATABLE SUBMISSION ==========

                # 1. Validate plating code using the resolved plating_color_internal
                plating_obj_code = plating_colors_internal.get(plating_color_internal)
                if not plating_obj_code:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Plating color internal code '{plating_color_internal}' not available in Master Data.")
                    continue

                # 2. Validate plating color from input
                plating_color_obj = plating_colors_name.get(plating_colour)
                if not plating_color_obj:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Plating Colour '{plating_colour}' not available in Master Data. Available: {_color_hint}")
                    continue

                # 3. Cross-validation: Check if resolved plating_color_internal matches with the plating color from Excel
                if plating_obj_code.pk != plating_color_obj.pk:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Plating color mismatch: Stock code '{plating_stock_no}' resolves to '{plating_obj_code.plating_color}' but Excel shows '{plating_colour}'.")
                    continue

                # 4. Validate polish code
                polish_obj = polish_types.get(polish_code)
                if not polish_obj:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Polish code '{polish_code}' not available in Master Data. Available: {_polish_hint}")
                    continue

                # 5. Version code validation
                version_obj = versions.get(version_code)
                if not version_obj:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Version code '{version_code}' not available in Master Data. Available: {_version_hint}")
                    continue

                # 6. Validate category
                category_obj = categories.get(category)
                if not category_obj:
                    failure_count += 1
                    failed_rows.append(f"Row {idx}: ❌ Category '{category}' not available in Master Data. Available: {_cat_hint}")
                    continue

                # 7. Validate source format - handle both underscore and non-underscore formats
                vendor_obj = None
                location_obj = None

                if "_" in source:
                    # Format: Vendor_Location (e.g., Titan_CPSE)
                    vendor_name, loc_name = source.split("_", 1)
                    
                    # Validate vendor by vendor_name field
                    vendor_obj = vendors.get(vendor_name)
                    if not vendor_obj:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Vendor '{vendor_name}' not available in Master Data. Available: {_vendor_hint}")
                        continue

                    # Validate location
                    location_obj = locations.get(loc_name)
                    if not location_obj:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Location '{loc_name}' not available in Master Data. Available: {_loc_hint}")
                        continue
                else:
                    # Format: Location only (e.g., CPSE, Mumbai, Delhi)
                    location_obj = locations.get(source)
                    if not location_obj:
                        failure_count += 1
                        failed_rows.append(f"Row {idx}: ❌ Location '{source}' not available in Master Data. Available: {_loc_hint}")
                        continue
                    vendor_obj = None

                # Generate batch ID using pre-computed timestamp
                batch_id = f"BATCH-{upload_ts}-{idx}"

                # Collect data for bulk create
                obj_data = {
                    'batch_id': batch_id,
                    'model_stock_no': model_stock,
                    'plating_color': plating_obj_code.plating_color,  # Use resolved plating color
                    'vendor_internal': vendor_obj.vendor_internal if vendor_obj else None,  # Handle None case
                    'location': location_obj,  # Can be None for vendor-only format
                    'tray_capacity': model_stock.tray_capacity if model_stock else None,
                    'tray_type': model_stock.tray_type.tray_type if model_stock and model_stock.tray_type else None,
                    'ep_bath_type': model_stock.ep_bath_type if model_stock else None,
                    'total_batch_quantity': input_qty,
                    'version': version_obj if version_obj else None,
                    'polish_finish': polish_obj if polish_obj else None,
                    'category': category_obj,  # Save category object instead of string
                    'plating_stk_no': plating_stock_no,           # <-- Save Plating Stk No
                    'polishing_stk_no': polishing_stock_no,  # <-- Save Polishing Stk No   
                }
                objects_to_create.append(ModelMasterCreation(**obj_data))
                success_count += 1

            # Bulk create all valid objects in a single transaction
            if objects_to_create:
                # Set sequence numbers
                last_sequence = ModelMasterCreation.objects.aggregate(max_seq=models.Max('sequence_number'))['max_seq'] or 0
                for i, obj in enumerate(objects_to_create, start=1):
                    obj.sequence_number = last_sequence + i
                with transaction.atomic():
                    ModelMasterCreation.objects.bulk_create(objects_to_create)
                print(f"✅ Bulk created {len(objects_to_create)} records")

            # Return results with enhanced error messages
            if success_count > 0 and failure_count == 0:
                messages.success(request, f"✅ {success_count} row(s) uploaded successfully.")
            elif success_count > 0 and failure_count > 0:
                error_msg = f"⚠️ Partial Success: {success_count} succeeded, {failure_count} failed.\n\n"
                error_msg += "\n".join(failed_rows)
                messages.warning(request, mark_safe(error_msg.replace('\n', '<br>')))
            elif success_count == 0 and failure_count > 0:
                error_msg = f"❌ Upload Failed: All {failure_count} row(s) failed.\n\n"
                error_msg += "\n".join(failed_rows)
                messages.error(request, mark_safe(error_msg.replace('\n', '<br>')))


            return Response({'master_data': ModelMasterCreation.objects.none()})

        except Exception as e:
            logger.error(f"❌ Error processing file: {str(e)}", exc_info=True)
            messages.error(request, "Unable to process the request. Please verify the submitted data and try again.")
            return Response({'master_data': ModelMasterCreation.objects.none()}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        finally:
            if wb:
                wb.close()

class DPBulkUploadPreviewView(APIView):
    """
    Enhanced preview view with detailed column validation
    """
    parser_classes = [MultiPartParser]

    def validate_excel_columns(self, sheet):
        """
        Validate Excel file columns to ensure they match expected format
        Returns: (is_valid, error_message, header_row)
        """
        try:
            # Expected column names in exact order
            expected_columns = [
                'S.No',
                'Plating Stk No', 
                'Polishing Stk No',
                'Plating Colour',
                'Category',
                'Input Qty',
                'Source'
            ]
            
            # Get the first row (header row)
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            
            if not header_row:
                return False, "❌ Excel file appears to be empty or corrupted.", None
            
            # Convert header values to strings and strip whitespace
            actual_columns = [str(cell).strip() if cell is not None else '' for cell in header_row]
            
            # Remove empty columns from the end
            while actual_columns and actual_columns[-1] == '':
                actual_columns.pop()
            
            # Check if we have the minimum required number of columns
            if len(actual_columns) < len(expected_columns):
                missing_count = len(expected_columns) - len(actual_columns)
                missing_columns = expected_columns[len(actual_columns):]
                return False, f"❌ Missing {missing_count} column(s): {', '.join(missing_columns)}. Expected {len(expected_columns)} columns but found {len(actual_columns)}.", actual_columns
            
            # Check for exact column name matches
            mismatched_columns = []
            for i, (expected, actual) in enumerate(zip(expected_columns, actual_columns)):
                if expected != actual:
                    mismatched_columns.append(f"Column {i+1}: Expected '{expected}' but found '{actual}'")
            
            if mismatched_columns:
                error_msg = "❌ Column Name Mismatch:\n" + "\n".join(mismatched_columns)
                error_msg += f"\n\nExpected format: {' | '.join(expected_columns)}"
                error_msg += f"\nActual format: {' | '.join(actual_columns[:len(expected_columns)])}"
                return False, error_msg, actual_columns
            
            # Check for extra columns (warn but don't fail)
            if len(actual_columns) > len(expected_columns):
                extra_columns = actual_columns[len(expected_columns):]
                print(f"⚠️ Warning: Found {len(extra_columns)} extra column(s) that will be ignored: {', '.join(extra_columns)}")
            
            return True, "✅ Column validation passed", actual_columns
            
        except Exception as e:
            return False, f"❌ Error reading Excel file headers: {str(e)}", None

    def post(self, request, *args, **kwargs):
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'success': False, 'error': 'No file uploaded.'}, status=400)
        
        if not uploaded_file.name.endswith(('.xls', '.xlsx')):
            return Response({
                'success': False, 
                'error': f"❌ Only Excel files are allowed. '{uploaded_file.name}' is not valid."
            }, status=400)
        
        try:
            wb = None
            wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
            sheet = wb.active
            if not sheet:
                return Response({
                    'success': False,
                    'error': '❌ Could not read the Excel sheet.'
                }, status=400)

            # ========== ENHANCED: USE DETAILED COLUMN VALIDATION ==========
            is_valid, error_message, actual_columns = self.validate_excel_columns(sheet)
            if not is_valid:
                return Response({
                    'success': False,
                    'error': error_message
                }, status=400)
            
            print("✅ Excel column validation passed for preview")
            # ========== END ENHANCED COLUMN VALIDATION ==========
            
            data = []
            total_rows_processed = 0
            skipped_rows = 0
            
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                total_rows_processed += 1
                if not any(row):
                    skipped_rows += 1
                    continue
                
                # Check for minimum required columns
                if len(row) < 7:
                    skipped_rows += 1
                    continue  # Skip incomplete rows in preview
                
                data.append({
                    'S.No': row[0] or '',
                    'Plating Stk No': row[1] or '',
                    'Polishing Stk No': row[2] or '',
                    'Plating Colour': row[3] or '',
                    'Category': row[4] or '',
                    'Input Qty': row[5] or '',
                    'Source': row[6] or '',
                })
            
            return Response({'success': True, 'data': data})

        except Exception as e:
            return Response({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)
        finally:
            if wb:
                wb.close()


# Alternative: Enhanced error messages with more specific column details
class DPBulkUploadPreviewViewEnhanced(APIView):
    """
    Even more enhanced preview view with ultra-specific column validation
    """
    parser_classes = [MultiPartParser]

    def validate_excel_columns_enhanced(self, sheet):
        """
        Enhanced validation with even more specific error messages
        """
        try:
            # Expected column names with their positions
            expected_columns = {
                1: 'S.No',
                2: 'Plating Stk No', 
                3: 'Polishing Stk No',
                4: 'Plating Colour',
                5: 'Category',
                6: 'Input Qty',
                7: 'Source'
            }
            
            # Get the first row (header row)
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            
            if not header_row:
                return False, "❌ Excel file appears to be empty or corrupted.", None
            
            # Convert header values to strings and strip whitespace
            actual_columns = [str(cell).strip() if cell is not None else '' for cell in header_row]
            
            # Remove empty columns from the end
            while actual_columns and actual_columns[-1] == '':
                actual_columns.pop()
            
            # Check if we have the minimum required number of columns
            if len(actual_columns) < len(expected_columns):
                missing_positions = list(range(len(actual_columns) + 1, len(expected_columns) + 1))
                missing_details = [f"Column {pos} should be '{expected_columns[pos]}'" for pos in missing_positions]
                return False, f"❌ Missing Required Columns:\n" + "\n".join(missing_details) + f"\n\nFound {len(actual_columns)} columns, expected {len(expected_columns)}.", actual_columns
            
            # Check each column position for exact matches
            specific_errors = []
            for pos in range(1, len(expected_columns) + 1):
                expected_name = expected_columns[pos]
                actual_name = actual_columns[pos - 1] if pos <= len(actual_columns) else ''
                
                if expected_name != actual_name:
                    specific_errors.append(
                        f"Column {pos} Name Mismatch: Expected '{expected_name}' but found '{actual_name}'"
                    )
            
            if specific_errors:
                error_msg = "❌ Column Name Mismatch Detected:\n\n" + "\n".join(specific_errors)
                error_msg += f"\n\n📋 Required Column Format:"
                for pos, name in expected_columns.items():
                    status = "✅" if pos <= len(actual_columns) and actual_columns[pos-1] == name else "❌"
                    current = actual_columns[pos-1] if pos <= len(actual_columns) else "MISSING"
                    error_msg += f"\n   Column {pos}: {name} {status} (Current: {current})"
                
                return False, error_msg, actual_columns
            
            return True, "✅ All column names validated successfully", actual_columns
            
        except Exception as e:
            return False, f"❌ Error reading Excel file headers: {str(e)}", None

    def post(self, request, *args, **kwargs):
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return Response({'success': False, 'error': 'No file uploaded.'}, status=400)
        
        if not uploaded_file.name.endswith(('.xls', '.xlsx')):
            return Response({
                'success': False, 
                'error': f"❌ Only Excel files are allowed. '{uploaded_file.name}' is not valid."
            }, status=400)
        
        try:
            wb = None
            wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
            sheet = wb.active
            if not sheet:
                return Response({
                    'success': False,
                    'error': '❌ Could not read the Excel sheet.'
                }, status=400)

            # Use enhanced column validation
            is_valid, error_message, actual_columns = self.validate_excel_columns_enhanced(sheet)
            if not is_valid:
                return Response({
                    'success': False,
                    'error': error_message
                }, status=400)
            
            # Rest of the processing logic remains the same...
            data = []
            for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                if not any(row):
                    continue
                
                if len(row) < 7:
                    continue
                
                data.append({
                    'S.No': row[0] or '',
                    'Plating Stk No': row[1] or '',
                    'Polishing Stk No': row[2] or '',
                    'Plating Colour': row[3] or '',
                    'Category': row[4] or '',
                    'Input Qty': row[5] or '',
                    'Source': row[6] or '',
                })
                
                if len(data) >= 100:
                    break
            
            return Response({'success': True, 'data': data})

        except Exception as e:
            return Response({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)
        finally:
            if wb:
                wb.close()


@method_decorator(csrf_exempt, name='dispatch')
class GetPlatingColourAPIView(APIView):
    """
    API endpoint to fetch plating colour based on color code from Plating Stk No
    
    POST /dayplanning/get_plating_colour/
    {
        "color_code": "S",
        "plating_stock_no": "1805SAB02"
    }
    
    Response:
    {
        "success": true,
        "plating_colour": "Silver",
        "color_code": "S"
    }
    """
    
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            color_code = data.get('color_code', '').strip().upper()
            plating_stock_no = data.get('plating_stock_no', '').strip()
            
            if not color_code:
                return JsonResponse({
                    'success': False,
                    'error': 'Color code is required'
                }, status=400)
            
            # Handle ambiguous color codes (with "/" identifier)
            plating_color_internal = color_code
            if "/" in plating_stock_no:
                # Extract the part after "/" for ambiguous resolution
                additional_identifier = plating_stock_no.split("/")[1]
                plating_color_internal = additional_identifier
                print(f"🔍 Using additional identifier for ambiguous color: {additional_identifier}")
            
            # Try to find plating color by internal code
            plating_color_obj = Plating_Color.objects.filter(
                plating_color_internal=plating_color_internal
            ).first()
            
            if plating_color_obj:
                return JsonResponse({
                    'success': True,
                    'plating_colour': plating_color_obj.plating_color,
                    'color_code': color_code,
                    'plating_color_internal': plating_color_internal
                })
            else:
                # Get available colors for suggestion
                available_colors = list(Plating_Color.objects.values_list(
                    'plating_color_internal', 'plating_color'
                )[:10])
                
                return JsonResponse({
                    'success': False,
                    'error': f"Plating color internal code '{plating_color_internal}' not found in Master Data.",
                    'available_colors': [
                        {'code': code, 'color': color} 
                        for code, color in available_colors
                    ]
                }, status=404)
                
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            logger.error(f"❌ Error in GetPlatingColourAPIView: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': 'Unable to process the request. Please verify the submitted data and try again.'
            }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class GetCategoriesAPIView(APIView):
    """
    API endpoint to fetch all categories for dropdown
    
    GET /dayplanning/get_categories/
    
    Response:
    {
        "success": true,
        "categories": [
            {"category_name": "Category A"},
            {"category_name": "Category B"}
        ]
    }
    """
    
    def get(self, request):
        try:
            # Import the Category model (you'll need to add this import at the top of views.py)
            from modelmasterapp.models import Category
            
            categories = list(Category.objects.values('category_name').order_by('category_name'))
            
            return JsonResponse({
                'success': True,
                'categories': categories,
                'count': len(categories)
            })
            
        except Exception as e:
            logger.error(f"❌ Error in GetCategoriesAPIView: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': 'Unable to process the request. Please verify the submitted data and try again.'
            }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class GetLocationsAPIView(APIView):
    """
    API endpoint to fetch all locations for dropdown
    
    GET /dayplanning/get_locations/
    
    Response:
    {
        "success": true,
        "locations": [
            {"location_name": "Location A"},
            {"location_name": "Location B"}
        ]
    }
    """
    
    def get(self, request):
        try:
            # Location model should already be imported
            locations = list(Location.objects.values('location_name').order_by('location_name'))
            
            return JsonResponse({
                'success': True,
                'locations': locations,
                'count': len(locations)
            })
            
        except Exception as e:
            logger.error(f"❌ Error in GetLocationsAPIView: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': 'Unable to process the request. Please verify the submitted data and try again.'
            }, status=500) 


class DayPlanningPickTableAPIView(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'Day_Planning/DP_PickTable.html'
    permission_classes = [IsAuthenticated]
 
    def get(self, request, *args, **kwargs):
        user = request.user
        with time_stage(request, 'DP_PERMISSION_CHECK'):
            # is_admin = user.groups.filter(name='Admin').exists() if user.is_authenticated else False
            is_admin = is_admin_user(user)
            module_name = "DP Pick Table"
            visible_headings = get_visible_headings_for_user(user, module_name)
            # Only include headings that are True (checked)
            allowed_headings = [h for h, v in visible_headings.items() if v]

        # Handle sorting parameters
        sort = request.GET.get('sort')
        order = request.GET.get('order', 'asc')  # Default to ascending
        
        # Field mapping for proper model field references
        sort_field_mapping = {
            'serial_number': 'batch_id',  # Use batch_id for serial number sorting
            'date_time': 'date_time',
            'plating_stk_no': 'plating_stk_no',
            'polishing_stk_no': 'polishing_stk_no',
            'plating_color': 'plating_color',
            'category': 'category',
            'polish_finish': 'polish_finish',
            'version': 'version__version_name',
            'tray_capacity': 'tray_capacity',
            'vendor_location': 'vendor_internal',  # Simplified to vendor field
            'no_of_trays': 'no_of_trays',
            'total_batch_quantity': 'total_batch_quantity',
            'dp_pick_remarks': 'dp_pick_remarks'
        }

        dp_data_fetch = time_stage(request, 'DP_DATA_FETCH')
        dp_data_fetch.__enter__()
        # Subqueries for annotations
        last_process_module_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('last_process_module')[:1]
        next_process_module_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('next_process_module')[:1]
        accepted_Ip_stock_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk'),
        ).values('accepted_Ip_stock')[:1]   
 
        # ✅ ENHANCED: Add subquery to check if tray scanning was started but not completed
        tray_scan_status_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('tray_scan_status')[:1]
 
        # Build queryset
        queryset = ModelMasterCreation.objects.filter(
            total_batch_quantity__gt=0,
            Moved_to_D_Picker=False  
        )
 
        filter_fields = [
            'batch_id', 'plating_color', 'polish_finish', 'version__version_name',
            'vendor_internal', 'location__location_name', 'tray_type', 'category',
            'tray_scan_status', 'model_stock_no__model_no', 'plating_stk_no', 'polishing_stk_no'
        ]
        for field in filter_fields:
            value = request.GET.get(field)
            if value:
                # For related fields, use __icontains for partial match
                lookup = f"{field}__icontains"
                queryset = queryset.filter(**{lookup: value})
        # --- ANNOTATE with subqueries ---
        last_process_module_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('last_process_module')[:1]
        next_process_module_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('next_process_module')[:1]
        accepted_Ip_stock_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk'),
        ).values('accepted_Ip_stock')[:1]
        tray_scan_status_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('tray_scan_status')[:1]

        queryset = queryset.annotate(
            last_process_module=Subquery(last_process_module_subquery),
            next_process_module=Subquery(next_process_module_subquery),
            accepted_Ip_stock=Subquery(accepted_Ip_stock_subquery),
            tray_scan_status=Subquery(tray_scan_status_subquery),
        )
        
        # Apply sorting if requested
        if sort and sort in sort_field_mapping:
            field = sort_field_mapping[sort]
            if order == 'desc':
                field = '-' + field
            queryset = queryset.order_by(field)
        else:
            queryset = queryset.order_by('-date_time', 'batch_id')  # Default sorting
            
        # Pagination
        page_number = request.GET.get('page', 1)
        paginator = Paginator(queryset, 10)
        page_obj = paginator.get_page(page_number)

        # Convert page_obj to list of dicts
        master_data = list(page_obj.object_list.values(
            'batch_id',
            'lot_id',
            'date_time',
            'model_stock_no__model_no',
            'plating_color',
            'polish_finish',
            'version__version_name',
            'version__version_internal',
            'vendor_internal',
            'location__location_name',
            'no_of_trays',
            'tray_type',
            'total_batch_quantity',
            'tray_capacity',
            'Moved_to_D_Picker',
            'last_process_module',  # <-- now valid
            'next_process_module',
            'Draft_Saved',
            'dp_pick_remarks',
            'top_tray_qty_verified',
            'plating_stk_no',
            'polishing_stk_no',
            'category',
            'hold_lot',
            'release_lot',
            'holding_reason',
            'release_reason',
            'accepted_Ip_stock',
            'tray_scan_status',
        ))
        dp_data_fetch.__exit__(None, None, None)

        dp_processing = time_stage(request, 'DP_PROCESSING')
        dp_processing.__enter__()
        # ✅ PERF: Batch-fetch per-batch data ONCE for the whole page instead of
        # issuing separate queries inside the row loop (was 3 queries per row →
        # N+1). Behaviour is identical; only the number of DB round-trips changes.
        page_batch_ids = [d['batch_id'] for d in master_data]

        # Batch ids that have a zero-quantity top tray (replaces per-row .exists()).
        zero_top_tray_batch_ids = set(
            TrayId.objects.filter(
                batch_id__batch_id__in=page_batch_ids,
                top_tray=True,
                tray_quantity=0,
            ).values_list('batch_id__batch_id', flat=True)
        )

        # Model images per batch (replaces per-row ModelMasterCreation fetch +
        # images.all()). One query for the rows, one for the M2M via prefetch.
        images_by_batch = {}
        for mmc in ModelMasterCreation.objects.filter(
            batch_id__in=page_batch_ids
        ).prefetch_related('images'):
            urls = [
                img.master_image.url
                for img in mmc.images.all()
                if getattr(img, 'master_image', None)
            ]
            images_by_batch[mmc.batch_id] = urls

        # Helper: normalize tray type string to pre-jig category and capacity
        def _get_prejig_tray(tray_type_str):
            tt = (tray_type_str or '').upper()
            if 'JUMBO' in tt or tt in ('JR', 'JD', 'JB', 'JL'):
                return 'JB', 12
            return 'NB', 16

        # Calculate no_of_trays dynamically and determine needs_top_tray_scan
        for data in master_data:
            total_batch_quantity = data.get('total_batch_quantity', 0)
            # Map tray type and capacity to pre-jig values (NB=16, JB=12)
            prejig_type, prejig_cap = _get_prejig_tray(data.get('tray_type'))
            data['tray_type'] = prejig_type
            data['tray_capacity'] = prejig_cap
            tray_capacity = prejig_cap
            data['vendor_location'] = f"{data.get('vendor_internal', '')}_{data.get('location__location_name', '')}"
 
            # ✅ ENHANCED: Determine if this lot needs top tray scan
            tray_scan_status = data.get('tray_scan_status', False)
            moved_to_d_picker = data.get('Moved_to_D_Picker', False)
            draft_saved = data.get('Draft_Saved', False)
            # ✅ ENHANCED: Check if there are any trays with quantity 0 in the first position
            # (precomputed once above — no per-row query).
            has_zero_top_tray = data['batch_id'] in zero_top_tray_batch_ids
            # ✅ ENHANCED: Multiple conditions for needing top tray scan
            data['needs_top_tray_scan'] = bool(
                (tray_scan_status and not moved_to_d_picker and not draft_saved) or  # Original condition, exclude draft
                (has_zero_top_tray and not draft_saved)  # New condition: zero quantity top tray exists, exclude draft
            )
            if tray_capacity > 0:
                no_of_trays = math.ceil(total_batch_quantity / tray_capacity)
                data['no_of_trays'] = no_of_trays
                tray_qty_list = []
                remainder = total_batch_quantity % tray_capacity
                if no_of_trays == 1:
                    tray_qty_list = [total_batch_quantity]
                elif no_of_trays > 1:
                    if remainder != 0:
                        tray_qty_list.append(remainder)
                        for _ in range(1, no_of_trays):
                            tray_qty_list.append(tray_capacity)
                    else:
                        for _ in range(no_of_trays):
                            tray_qty_list.append(tray_capacity)
                data['tray_qty_list'] = tray_qty_list
            else:
                data['no_of_trays'] = 0
                data['tray_qty_list'] = []
            # Add model images (precomputed once above — no per-row query).
            images = list(images_by_batch.get(data['batch_id'], []))
            if not images:
                from django.templatetags.static import static
                images = [static('assets/images/imagePlaceholder.jpg')]
            data['model_images'] = images
        dp_processing.__exit__(None, None, None)

        # ✅ ENHANCED: Define correct column order for table headers
        correct_column_order = [
            'S.No',
            'Last Updated',
            'Plating Stk No',
            'Action',
            'No of Trays',
            'Input Qty',
            'Process Status',
            'Lot Status',
            'Current Stage',
            'Polishing Stk No',
            'Plating Color',
            'Category',
            'Polish Finish',
            'Version',
            'Tray Cate-Capacity',
            'Source',
            'Remarks',
        ]
        
        # ✅ ENHANCED: Create display headings map for better presentation
        display_headings_map = {
            'S.No': 'S.No',
            'Last Updated': 'Last Updated',
            'Plating Stk No': 'Plating Stock Number',
            'Action': 'Action',
            'No of Trays': 'No of Trays',
            'Input Qty': 'Input Qty',
            'Process Status': 'Process Status',
            'Lot Status': 'Lot Status',
            'Current Stage': 'Current Stage',
            'Polishing Stk No': 'Polishing Stk No',
            'Plating Color': 'Plating Color',
            'Category': 'Category',
            'Polish Finish': 'Polish Finish',
            'Version': 'Version',
            'Tray Cate-Capacity': 'Tray Cate-Capacity',
            'Source': 'Source',
            'Remarks': 'Remarks',
        }
        
        # ✅ ENHANCED: Reorder visible_headings to match correct column order
        # Keep only columns that are in the correct order and preserve visibility settings
        ordered_visible_headings = {}
        for col in correct_column_order:
            if col in visible_headings:
                ordered_visible_headings[col] = visible_headings[col]
            else:
                # If column not in visible_headings, default to visible for admin
                ordered_visible_headings[col] = is_admin
        
        # Update allowed_headings to match the new order
        ordered_allowed_headings = [h for h in correct_column_order if ordered_visible_headings.get(h, False)]
        
        context = {
            'master_data': master_data,
            'page_obj': page_obj,
            'paginator': paginator,
            'user': user,
            'is_admin': is_admin,
            'visible_headings': ordered_visible_headings,
            'allowed_headings': ordered_allowed_headings,
            'display_headings_map': display_headings_map,
        }
        return Response(context, template_name=self.template_name)

@method_decorator(csrf_exempt, name='dispatch')
class SaveHoldUnholdReasonAPIView(APIView):
    """
    POST with:
    {
        "batch_id": "BATCH-20240618123456-1",
        "remark": "Reason text",
        "action": "hold",  # or "unhold"
        "previous_status": "yet_to_start"  # optional, for restoring status
    }
    """
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            remark = data.get('remark', '').strip()
            action = data.get('action', '').strip().lower()
            previous_status = data.get('previous_status', '').strip()
            user = request.user if request.user.is_authenticated else None
            
            # Print user and action to command prompt
            print(f'[ROW HOLD] User "{user.username if user else "Anonymous"}" is performing "{action}" on BatchID: {batch_id} with remark: "{remark}"')

            if not batch_id or not remark or action not in ['hold', 'unhold']:
                return JsonResponse({'success': False, 'error': 'Missing or invalid parameters.'}, status=400)

            obj = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found.'}, status=404)

            if action == 'hold':
                # Store current lot status before holding
                if obj.Moved_to_D_Picker:
                    obj.previous_lot_status = 'yet_to_release'
                elif obj.Draft_Saved:
                    obj.previous_lot_status = 'draft'

                else:
                    obj.previous_lot_status = 'yet_to_start'
                
                obj.holding_reason = remark
                obj.hold_lot = True
                obj.release_reason = ''
                obj.release_lot = False
                
            elif action == 'unhold':
                obj.release_reason = remark
                obj.hold_lot = False
                obj.release_lot = True
                # previous_lot_status is preserved for frontend to restore status

            obj.save(update_fields=[
                'holding_reason', 'release_reason', 'hold_lot', 'release_lot', 'previous_lot_status'
            ])
            
            return JsonResponse({
                'success': True, 
                'message': 'Reason saved.',
                'previous_status': obj.previous_lot_status
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class TrayIdScanAPIView(APIView):
    """
    Enhanced TrayId Scan API - ONLY ALLOW PRE-EXISTING TRAYS (no new tray creation)
    """

    def post(self, request):
        try:
            if hasattr(request, 'data'):
                data = request.data
            else:
                data = json.loads(request.body.decode('utf-8'))

            batch_id = data.get('batch_id')
            trays = data.get('trays', [])
            user = request.user if request.user.is_authenticated else None

            lot_id = data.get('lot_id') or self.generate_new_lot_id()

            if not batch_id or not trays:
                return JsonResponse({'success': False, 'error': 'Missing required fields.'}, status=400)

            batch_instance = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not batch_instance:
                return JsonResponse({'success': False, 'error': 'Invalid batch_id.'}, status=400)

            # ✅ ENHANCED: Pre-validate all tray IDs - MUST EXIST IN SYSTEM
            tray_type_errors = []
            already_scanned_errors = []
            duplicate_tray_ids = []
            tray_not_in_system_errors = []  # ✅ NEW: Track trays not in system
            
            for i, tray in enumerate(trays):
                tray_id = tray.get('tray_id', '').strip()
                if not tray_id:
                    continue
                
                # Allow partial typing: only show error when first two chars are present and not JB/NB
                prefix = tray_id.upper()
                first2 = prefix[:2]
                if len(first2) == 2 and first2 not in ('JB', 'NB'):
                    tray_not_in_system_errors.append({
                        'tray_id': tray_id,
                        'position': i + 1,
                        'prefix_error': True,
                    })
                    continue
                
                # ✅ NEW: First check if tray exists in TrayId table
                existing_tray = TrayId.objects.filter(tray_id=tray_id).first()
                
                if not existing_tray:
                    # ✅ NEW: Reject trays that don't exist in system
                    tray_not_in_system_errors.append({
                        'tray_id': tray_id,
                        'position': i + 1,
                        'error': f'Tray ID "{tray_id}" not found in system. Only pre-configured trays are allowed.'
                    })
                    continue
                
                # ✅ Check if tray is already scanned (and not delinked)
                if existing_tray.scanned and not existing_tray.delink_tray:
                    already_scanned_errors.append({
                        'tray_id': tray_id,
                        'position': i + 1,
                        'batch_info': existing_tray.batch_id.batch_id if existing_tray.batch_id else 'Unknown batch',
                        'scan_date': existing_tray.date.strftime('%d-%m-%Y %H:%M') if existing_tray.date else 'Unknown date'
                    })
                    continue
                
                # Check if tray belongs to different batch (and not delinked)
                if (existing_tray.batch_id and 
                    existing_tray.batch_id != batch_instance and 
                    not existing_tray.delink_tray):
                    duplicate_tray_ids.append(tray_id)
                    continue
                
                # Validate tray type for existing trays
                validation_helper = TrayIdUniqueCheckAPIView()
                validation_result = validation_helper.validate_tray_type_compatibility(
                    existing_tray, batch_id
                )
                
                if not validation_result['compatible']:
                    tray_type_errors.append({
                        'tray_id': tray_id,
                        'position': i + 1,
                        'error': validation_result['error'],
                        'batch_tray_type': validation_result['batch_tray_type'],
                        'scanned_tray_type': validation_result['scanned_tray_type']
                    })

            # ✅ NEW: Return tray not in system errors
            if tray_not_in_system_errors:
                error_messages = []
                for error in tray_not_in_system_errors:
                    error_messages.append(
                        f"Position {error['position']}: {error['tray_id']} - Not found in system"
                    )
                
                return JsonResponse({
                    'success': False,
                    'error': 'Some trays are not in the system',
                    'tray_not_in_system_errors': tray_not_in_system_errors,
                    'error_details': error_messages
                }, status=400)

            # Return already scanned errors if any
            if already_scanned_errors:
                error_messages = []
                for error in already_scanned_errors:
                    error_messages.append(
                        f"Position {error['position']}: {error['tray_id']} - Already scanned on {error['scan_date']}"
                    )
                
                return JsonResponse({
                    'success': False,
                    'error': 'Some trays are already scanned',
                    'already_scanned_errors': already_scanned_errors,
                    'error_details': error_messages
                }, status=400)

            # Return tray type errors if any
            if tray_type_errors:
                error_messages = []
                for error in tray_type_errors:
                    error_messages.append(f"Position {error['position']}: {error['error']}")
                
                return JsonResponse({
                    'success': False,
                    'error': 'Tray type validation failed',
                    'tray_type_errors': tray_type_errors,
                    'error_details': error_messages
                }, status=400)

            # Return duplicate errors if any
            if duplicate_tray_ids:
                return JsonResponse({
                    'success': False,
                    'error': f'Duplicate Tray ID(s) found: {", ".join(duplicate_tray_ids)}',
                    'duplicate_tray_ids': duplicate_tray_ids
                }, status=400)

            # Check if top tray (first tray) has quantity 0
            top_tray_qty_zero = False
            if trays and len(trays) > 0:
                first_tray_qty = int(trays[0].get('tray_quantity', 0))
                if first_tray_qty == 0:
                    top_tray_qty_zero = True
                    print(f"🔍 Top tray quantity is 0 - special handling required")

            # Fetch related fields from batch_instance
            model_stock_no = batch_instance.model_stock_no
            version = batch_instance.version
            total_batch_quantity = batch_instance.total_batch_quantity
            polish_finish_obj = batch_instance.polish_finish if isinstance(batch_instance.polish_finish, PolishFinishType) else PolishFinishType.objects.filter(polish_finish=batch_instance.polish_finish).first()
            plating_color = batch_instance.plating_color

            # In the TrayIdScanAPIView post method, find this section and update it:

            with transaction.atomic():
                # Calculate total quantity excluding delinked trays (qty = 0)
                active_total_quantity = sum(int(tray.get('tray_quantity', 0)) for tray in trays if int(tray.get('tray_quantity', 0)) > 0)
                
                # Save TotalStockModel with active quantity only
                total_stock_obj = TotalStockModel.objects.create(
                    batch_id=batch_instance,
                    model_stock_no=model_stock_no,
                    version=version,
                    total_stock=active_total_quantity,
                    polish_finish=polish_finish_obj,
                    plating_color=batch_instance.plating_color if isinstance(batch_instance.plating_color, Plating_Color) else Plating_Color.objects.filter(plating_color=batch_instance.plating_color).first() if batch_instance.plating_color else None,
                    lot_id=lot_id,
                    tray_scan_status=True,  # ✅ Mark as True to indicate tray scanning started
                    ip_draft_screening=False,  # ✅ Initialize IS draft flag
                    last_process_module="DayPlanning",
                    next_process_module="IP Screening",
                )

                # Process each tray - ONLY UPDATE EXISTING TRAYS
                for i, tray in enumerate(trays):
                    tray_id = tray.get('tray_id')
                    tray_quantity = tray.get('tray_quantity')
                    
                    if not tray_id or tray_quantity is None:
                        continue
                    
                    tray_quantity = int(tray_quantity)
                    
                    # Only set top_tray=True for first tray if its quantity > 0
                    if top_tray_qty_zero:
                        is_top_tray = False
                    else:
                        is_top_tray = (i == 0)
                    
                    # Check if tray should be delinked (quantity = 0)
                    is_delinked = (tray_quantity == 0)
                    delink_qty = None
                    
                    if is_delinked:
                        # Store original quantity from draft or previous value
                        draft_tray = DraftTrayId.objects.filter(
                            batch_id=batch_instance, 
                            tray_id=tray_id
                        ).first()
                        if draft_tray and draft_tray.delink_tray_qty:
                            delink_qty = draft_tray.delink_tray_qty
                        else:
                            delink_qty = str(batch_instance.tray_capacity or 0)
                    
                    # ✅ CHANGED: ONLY UPDATE EXISTING TRAYS (no new tray creation)
                    existing_tray = TrayId.objects.filter(tray_id=tray_id).first()
                    
                    if existing_tray:
                        # Update existing tray while preserving tray_type and tray_capacity
                        existing_tray.lot_id = lot_id
                        existing_tray.batch_id = batch_instance
                        existing_tray.tray_quantity = tray_quantity
                        existing_tray.user = user
                        existing_tray.delink_tray = is_delinked
                        existing_tray.delink_tray_qty = delink_qty
                        existing_tray.top_tray = is_top_tray
                        existing_tray.date = now()
                        existing_tray.scanned = True  # Set scanned status to True
                        existing_tray.new_tray = False  # <-- Set new_tray to False as requested
                        
                        from .models import DPTrayId_History  # Import at the top if not already

                        DPTrayId_History.objects.create(
                            lot_id=lot_id,
                            tray_id=tray_id,
                            tray_quantity=tray_quantity,
                            batch_id=batch_instance,
                            date=now(),
                            user=user,
                            top_tray=is_top_tray,
                            delink_tray=False,
                            rejected_tray=False,
                            new_tray=False,
                            tray_type=existing_tray.tray_type,
                            tray_capacity=existing_tray.tray_capacity,
                        )

                        # Don't update tray_type and tray_capacity - preserve admin settings
                        if is_delinked:
                            existing_tray.delink_tray = True
                            existing_tray.lot_id = None
                            existing_tray.batch_id = None
                            existing_tray.scanned = False
                            existing_tray.IP_tray_verified = False
                            existing_tray.top_tray = False
                            existing_tray.save(update_fields=[
                                'delink_tray', 'lot_id', 'batch_id', 'scanned', 'IP_tray_verified', 'top_tray'
                            ])
                        else:
                            # Don't update tray_type and tray_capacity - preserve admin settings
                            existing_tray.delink_tray = False
                            existing_tray.save()
                        print(f"✅ Updated existing tray: {tray_id} (Top Tray: {is_top_tray}, Type: {existing_tray.tray_type}, Scanned: True)")
                    else:
                        # ✅ CHANGED: This should never happen now since we validate existence above
                        print(f"❌ ERROR: Tray {tray_id} not found in system - this should have been caught in validation")
                        return JsonResponse({
                            'success': False, 
                            'error': f'Tray ID {tray_id} not found in system'
                        }, status=400)
                

                # ✅ ENHANCED: Handle Moved_to_D_Picker based on top tray quantity
                if top_tray_qty_zero:
                    # ✅ NEW: If top tray quantity is zero, keep as draft mode (Moved_to_D_Picker=False)
                    # This allows user to complete the "Set Top Tray" step later
                    print(f"🔍 Top tray quantity is zero - keeping Moved_to_D_Picker=False (draft mode)")
                    # Do not update Moved_to_D_Picker, it remains False
                else:
                    # ✅ NORMAL: Complete tray scanning process
                    ModelMasterCreation.objects.filter(batch_id=batch_id).update(Moved_to_D_Picker=True)
                    print(f"✅ Normal tray scanning completed - set Moved_to_D_Picker=True")

            # Return response based on top tray quantity
            if top_tray_qty_zero:
                return JsonResponse({
                    'success': True, 
                    'message': 'Tray scan saved! Please scan top tray ID.',
                    'top_tray_scan_required': True,
                    'batch_id': batch_id,
                    'keep_modal_open': True
                }, status=201)
            else:
                return JsonResponse({'success': True, 'message': 'Tray scan and stock saved!'}, status=201)
        except IntegrityError as e:
            error_message = str(e)
            if 'duplicate key value violates unique constraint' in error_message and 'tray_id' in error_message:
                import re
                tray_match = re.search(r'Key \(tray_id\)=\(([^)]+)\)', error_message)
                duplicate_tray = tray_match.group(1) if tray_match else 'unknown'
                return JsonResponse({
                    'success': False, 
                    'error': f'Tray ID {duplicate_tray} already exists in the system.',
                    'duplicate_tray_ids': [duplicate_tray]
                }, status=400)
            else:
                return JsonResponse({'success': False, 'error': 'Database integrity error: ' + str(e)}, status=400)
        except Exception as e:
            logger.error(f"❌ Error in TrayIdScanAPIView: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Unexpected error: ' + str(e)}, status=500)

    def generate_new_lot_id(self):
        timestamp = datetime.now().strftime("%d%m%Y%H%M%S")
        next_seq_no = 1
        # Iterate recent lots to find last sequential (non-UUID) lot ID
        for lot in TotalStockModel.objects.order_by('-id')[:20]:
            if lot.lot_id and lot.lot_id.startswith("LID"):
                try:
                    last_seq_no = int(lot.lot_id[-4:])
                    next_seq_no = last_seq_no + 1
                    break
                except ValueError:
                    continue
        seq_no = f"{next_seq_no:04d}"
        return f"LID{timestamp}{seq_no}"


# Add this new API view to your views.py file

@method_decorator(csrf_exempt, name='dispatch')
class ValidateTopTrayAPIView(APIView):
    """
    Enhanced API endpoint for validating tray ID for top tray selection - ONLY ALLOW EXISTING TRAYS
    """
    def post(self, request):
        try:
            if hasattr(request, 'data'):
                data = request.data
            else:
                data = json.loads(request.body.decode('utf-8'))

            batch_id = data.get('batch_id')
            tray_id = data.get('tray_id', '').strip()

            if not batch_id or not tray_id:
                return JsonResponse({
                    'success': False,
                    'valid': False,
                    'error': 'Missing batch ID or tray ID'
                }, status=400)

            # Check if batch exists
            batch_instance = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not batch_instance:
                return JsonResponse({
                    'success': False,
                    'valid': False,
                    'error': 'Batch not found'
                }, status=404)

            # Check if batch is already completed
            if batch_instance.Moved_to_D_Picker:
                return JsonResponse({
                    'success': True,
                    'valid': False,
                    'message': 'Batch already completed'
                })

            # ✅ NEW: First check if tray exists in TrayId table at all
            tray_in_system = TrayId.objects.filter(tray_id=tray_id).first()
            if not tray_in_system:
                return JsonResponse({
                    'success': True,
                    'valid': False,
                    'error': f'Tray ID "{tray_id}" not found in system. Only pre-configured trays are allowed.'
                })

            # Allow partial typing: only error when first two chars are present and not JB/NB
            prefix = tray_id.upper()
            first2 = prefix[:2]
            if len(first2) == 2 and first2 not in ('JB', 'NB'):
                return JsonResponse({
                    'success': True,
                    'valid': False,
                    'prefix_error': True,
                })

            # Check if tray exists in this specific batch (TrayId or DraftTrayId)
            tray = TrayId.objects.filter(
                tray_id=tray_id,
                batch_id=batch_instance
            ).first()

            draft_tray = DraftTrayId.objects.filter(
                tray_id=tray_id,
                batch_id=batch_instance
            ).first()

            if not tray and not draft_tray:
                return JsonResponse({
                    'success': True,
                    'valid': False,
                    'error': f'Tray ID "{tray_id}" not found in this batch'
                })

            # Use the tray from TrayId if exists, otherwise from DraftTrayId
            selected_tray = tray if tray else draft_tray

            # Check if the tray is delinked (only for TrayId trays)
            if tray and tray.delink_tray:
                return JsonResponse({
                    'success': True,
                    'valid': False,
                    'error': f'Tray ID "{tray_id}" is delinked and cannot be set as top tray'
                })

            # Check if the tray has quantity 0
            if selected_tray.tray_quantity <= 0:
                return JsonResponse({
                    'success': True,
                    'valid': False,
                    'error': f'Tray ID "{tray_id}" has zero quantity and cannot be set as top tray'
                })

            # Validate tray type compatibility (only for existing TrayId trays)
            if tray:
                validation_helper = TrayIdUniqueCheckAPIView()
                validation_result = validation_helper.validate_tray_type_compatibility(tray, batch_id)
                
                if not validation_result['compatible']:
                    return JsonResponse({
                        'success': True,
                        'valid': False,
                        'error': validation_result['error'],
                        'batch_tray_type': validation_result['batch_tray_type'],
                        'scanned_tray_type': validation_result['scanned_tray_type']
                    })

            # All checks passed: Valid tray for top tray selection
            return JsonResponse({
                'success': True,
                'valid': True,
                'message': f'Tray ID "{tray_id}" is valid for top tray selection'
            })

        except Exception as e:
            logger.error(f"❌ Error in ValidateTopTrayAPIView: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'valid': False,
                'error': 'Unable to process the request. Please verify the submitted data and try again.'
            }, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class TopTrayScanAPIView(APIView):
    """
    API endpoint to handle top tray scanning when original top tray quantity is 0
    Enhanced to complete the tray scanning process by setting Moved_to_D_Picker=True
    """
    def post(self, request):
        try:
            if hasattr(request, 'data'):
                data = request.data
            else:
                data = json.loads(request.body.decode('utf-8'))

            batch_id = data.get('batch_id')
            scanned_tray_id = data.get('scanned_tray_id', '').strip()
            user = request.user if request.user.is_authenticated else None

            if not batch_id or not scanned_tray_id:
                return JsonResponse({
                    'success': False, 
                    'error': 'Missing batch ID or tray ID.'
                }, status=400)

            # Allow partial typing: only error when first two chars are present and not JB/NB
            prefix = scanned_tray_id.upper()
            first2 = prefix[:2]
            if len(first2) == 2 and first2 not in ('JB', 'NB'):
                return JsonResponse({
                    'success': False,
                    'error': f'Tray ID "{scanned_tray_id}" is not allowed. Only JB and NB trays are permitted.'
                }, status=400)

            batch_instance = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not batch_instance:
                return JsonResponse({
                    'success': False, 
                    'error': 'Invalid batch ID.'
                }, status=404)

            # Check if batch is already completed
            if batch_instance.Moved_to_D_Picker:
                return JsonResponse({
                    'success': True,
                    'message': 'Batch already completed'
                })

            with transaction.atomic():
                # Check if the scanned tray ID exists in the same batch (TrayId or DraftTrayId)
                existing_tray = TrayId.objects.filter(
                    tray_id=scanned_tray_id,
                    batch_id=batch_instance
                ).first()

                draft_tray = DraftTrayId.objects.filter(
                    tray_id=scanned_tray_id,
                    batch_id=batch_instance
                ).first()

                if not existing_tray and not draft_tray:
                    return JsonResponse({
                        'success': False,
                        'error': f'Tray ID "{scanned_tray_id}" not found in this batch.'
                    }, status=400)

                # If found in draft, create TrayId entry
                if not existing_tray and draft_tray:
                    existing_tray = TrayId.objects.create(
                        tray_id=draft_tray.tray_id,
                        tray_quantity=draft_tray.tray_quantity,
                        batch_id=batch_instance,
                        position=draft_tray.position,
                        date=draft_tray.date,
                        # Set other default fields as needed
                    )
                    draft_tray.delete()  # Remove from draft

                # Check if this tray is delinked
                if existing_tray.delink_tray:
                    return JsonResponse({
                        'success': False,
                        'error': f'Tray ID "{scanned_tray_id}" is delinked and cannot be set as top tray.'
                    }, status=400)

                # ✅ UPDATED: Remove top_tray status from ALL trays in this batch first
                TrayId.objects.filter(
                    batch_id=batch_instance,
                    top_tray=True
                ).update(top_tray=False)

                # Set ONLY the scanned tray as the new top tray in TrayId
                existing_tray.top_tray = True
                existing_tray.save(update_fields=['top_tray'])

                # ✅ Also update DPTrayId_History for this batch and tray
                from .models import DPTrayId_History  # Import if not already at top
                DPTrayId_History.objects.filter(
                    batch_id=batch_instance,
                    tray_id=scanned_tray_id
                ).update(top_tray=True)

                # Optionally, set top_tray=False for all other trays in DPTrayId_History for this batch
                DPTrayId_History.objects.filter(
                    batch_id=batch_instance
                ).exclude(tray_id=scanned_tray_id).update(top_tray=False)

                # Complete the tray scanning process by setting Moved_to_D_Picker=True
                batch_instance.Moved_to_D_Picker = True
                batch_instance.save(update_fields=['Moved_to_D_Picker'])
                            
                print(f"✅ Set Moved_to_D_Picker=True for batch {batch_id} - tray scanning completed")

                return JsonResponse({
                    'success': True,
                    'message': f'Top tray set successfully: {scanned_tray_id}. Tray scanning completed!'
                })

        except Exception as e:
            logger.error(f"❌ Error in TopTrayScanAPIView: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False, 
                'error': 'Unable to process the request. Please verify the submitted data and try again.'
            }, status=500)
            

@method_decorator(csrf_exempt, name='dispatch')
class VerifyTopTrayQtyAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            verified_tray_qty = data.get('verified_tray_qty')
            
            print(f"🔍 VerifyTopTrayQty called with batch_id={batch_id}, verified_tray_qty={verified_tray_qty}")
            
            if not batch_id or verified_tray_qty is None:
                return JsonResponse({'success': False, 'error': 'Missing batch_id or verified_tray_qty'}, status=400)
            
            obj = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found'}, status=404)

            print(f"📋 Found batch: {obj.batch_id}")
            print(f"📊 Current values: total_batch_quantity={obj.total_batch_quantity}, initial_batch_quantity={obj.initial_batch_quantity}")

            # Store the original total_batch_quantity as initial_batch_quantity if not already set
            if not obj.initial_batch_quantity:
                obj.initial_batch_quantity = obj.total_batch_quantity
                print(f"💾 Set initial_batch_quantity = {obj.initial_batch_quantity}")

            # Update the verified quantity and verification status
            obj.verified_tray_qty = verified_tray_qty
            obj.total_batch_quantity = verified_tray_qty  # Update total to verified amount
            obj.top_tray_qty_verified = True
            obj.Draft_Saved = True
            
            # Save the changes
            obj.save(update_fields=[
                'top_tray_qty_verified', 
                'verified_tray_qty', 
                'total_batch_quantity', 
                'initial_batch_quantity',
                'Draft_Saved'
            ])
            
            print(f"✅ Successfully updated batch {batch_id}")
            print(f"📈 Final values: total_batch_quantity={obj.total_batch_quantity}, verified_tray_qty={obj.verified_tray_qty}, top_tray_qty_verified={obj.top_tray_qty_verified}")
            
            return JsonResponse({
                'success': True, 
                'message': 'Top tray quantity verified successfully.',
                'verified_tray_qty': obj.verified_tray_qty,
                'total_batch_quantity': obj.total_batch_quantity,
                'top_tray_qty_verified': obj.top_tray_qty_verified
            })
                     
        except Exception as e:
            logger.error(f"❌ Error in VerifyTopTrayQtyAPIView: {str(e)}", exc_info=True)
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class TrayIdListAPIView(APIView):
    def get(self, request):
        batch_id = request.GET.get('batch_id')
        if not batch_id:
            return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
        
        # ✅ UPDATED: Get top tray first, then other trays
        # First, get the top tray (top_tray=True)
        top_tray = TrayId.objects.filter(
            batch_id__batch_id=batch_id,
            delink_tray=False,
            top_tray=True
        ).first()
        
        # Then get all other trays (excluding the top tray)
        other_trays = TrayId.objects.filter(
            batch_id__batch_id=batch_id,
            delink_tray=False,
            top_tray=False
        ).order_by('id')
        
        data = []
        
        # Add top tray first if it exists
        if top_tray:
            data.append({
                's_no': 1,  # Always 1 for top tray
                'tray_id': top_tray.tray_id,
                'tray_quantity': top_tray.tray_quantity,
                'position': 0,  # Top position
                'is_top_tray': True
            })
        
        # Add other trays starting from position 2
        for idx, tray in enumerate(other_trays):
            data.append({
                's_no': idx + 2,  # Start from 2 since top tray is 1
                'tray_id': tray.tray_id,
                'tray_quantity': tray.tray_quantity,
                'position': idx + 1,  # Position after top tray
                'is_top_tray': False
            })
        
        return JsonResponse({'success': True, 'trays': data})
    
    
    
@method_decorator(csrf_exempt, name='dispatch')
class DraftTrayIdAPIView(APIView):
    def post(self, request):
        try:
            if hasattr(request, 'data'):
                data = request.data
            else:
                data = json.loads(request.body.decode('utf-8'))

            batch_id = data.get('batch_id')
            trays = data.get('trays', [])
            # Read top tray verification state and verified qty if sent
            top_tray_qty_verified = data.get('top_tray_qty_verified', None)
            verified_tray_qty = data.get('verified_tray_qty', None)
            user = request.user if request.user.is_authenticated else None

            if not batch_id or not trays:
                return JsonResponse({'success': False, 'error': 'Missing required fields.'}, status=400)

            batch_instance = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not batch_instance:
                return JsonResponse({'success': False, 'error': 'Invalid batch_id.'}, status=400)

            lot_id = data.get('lot_id') or f"DRAFT-{batch_id}"

            with transaction.atomic():
                # Always clear all existing draft entries for this batch before saving new ones
                DraftTrayId.objects.filter(batch_id=batch_instance).delete()

                # ✅ CRITICAL FIX: Calculate total quantity from drafted trays
                total_drafted_qty = 0

                for tray in trays:
                    tray_id = tray.get('tray_id', '').strip()
                    tray_quantity = tray.get('tray_quantity')
                    position = tray.get('position')
                    
                    if tray_quantity is None or position is None:
                        continue
                    
                    tray_quantity = int(tray_quantity)
                    is_delinked = (tray_quantity == 0)
                    delink_qty = None

                    # ✅ Add to total (including zero quantities)
                    total_drafted_qty += tray_quantity

                    DraftTrayId.objects.create(
                        lot_id=lot_id,
                        batch_id=batch_instance,
                        tray_id=tray_id,
                        tray_quantity=tray_quantity,
                        position=position,
                        user=user,
                        delink_tray=is_delinked,
                        delink_tray_qty=delink_qty
                    )
                    print(f"✅ Saved draft: Position {position}, Tray ID: '{tray_id}', Qty: {tray_quantity}, Delinked: {is_delinked}")

                # ✅ CRITICAL FIX: Update ModelMasterCreation total_batch_quantity with drafted sum
                print(f"📊 Draft total calculation: Old total_batch_quantity={batch_instance.total_batch_quantity}, New total_drafted_qty={total_drafted_qty}")
                batch_instance.total_batch_quantity = total_drafted_qty
                batch_instance.Draft_Saved = True
                
                # ✅ CRITICAL FIX: Create or update TotalStockModel to set tray_scan_status
                total_stock_obj, created = TotalStockModel.objects.get_or_create(
                    batch_id=batch_instance,
                    defaults={
                        'model_stock_no': batch_instance.model_stock_no,
                        'version': batch_instance.version,
                        'total_stock': total_drafted_qty,
                        'polish_finish': batch_instance.polish_finish if isinstance(batch_instance.polish_finish, PolishFinishType) else PolishFinishType.objects.filter(polish_finish=batch_instance.polish_finish).first(),
                        'plating_color': batch_instance.plating_color if isinstance(batch_instance.plating_color, Plating_Color) else Plating_Color.objects.filter(plating_color=batch_instance.plating_color).first() if batch_instance.plating_color else None,
                        'lot_id': lot_id,
                        'tray_scan_status': True,  # Boolean field, not string
                        'last_process_module': "DayPlanning",
                        'next_process_module': "IP Screening",
                    }
                )
                
                if not created:
                    # Update existing TotalStockModel
                    total_stock_obj.tray_scan_status = True
                    total_stock_obj.total_stock = total_drafted_qty
                    total_stock_obj.save(update_fields=['tray_scan_status', 'total_stock'])

                # If frontend sent top-tray verification state, persist it; otherwise set based on top tray quantity
                update_fields = ['Draft_Saved', 'total_batch_quantity']
                if top_tray_qty_verified is not None:
                    batch_instance.top_tray_qty_verified = True if str(top_tray_qty_verified) in ['True', 'true', True] else False
                    update_fields.append('top_tray_qty_verified')
                else:
                    # For draft, always set as verified to avoid showing top tray scan UI
                    batch_instance.top_tray_qty_verified = True
                    update_fields.append('top_tray_qty_verified')

                batch_instance.save(update_fields=update_fields)
                # Save the extra fields if any
                if any(x in update_fields for x in ['top_tray_qty_verified', 'verified_tray_qty']):
                    # Ensure we call save again for the additional fields included
                    batch_instance.save(update_fields=list(set(update_fields)))

                print(f"✅ Updated batch {batch_id}: total_batch_quantity={total_drafted_qty}, Draft_Saved=True")

            return JsonResponse({
                'success': True,
                'message': 'Draft saved!',
                'top_tray_qty_verified': batch_instance.top_tray_qty_verified,
                'verified_tray_qty': batch_instance.verified_tray_qty
            }, status=201)

        except IntegrityError as e:
            logger.error(f"❌ IntegrityError in DraftTrayIdAPIView: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False, 
                'error': 'Unable to process the request. Please verify the submitted data and try again.'
            }, status=400)
        except Exception as e:
            logger.error(f"❌ Error in DraftTrayIdAPIView: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Unexpected error: ' + str(e)}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class GlobalDraftedTraysAPIView(APIView):
    """
    Return trays that are genuinely active in Day Planning across all lots.
    Accepts optional batch_id to EXCLUDE the current batch's own trays so that
    re-opening a draft modal does not flag the batch's own trays as duplicates.

    Uses DPTrayId_History (the real DP transaction log) instead of TrayId master
    so that trays which have moved past Day Planning (accepted to IS/Brass QC etc.)
    are NOT reported as globally used.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from .models import DPTrayId_History
        current_batch_id = request.GET.get('batch_id', '').strip()

        # ── Active drafted trays (not yet submitted) ──────────────────────────
        draft_qs = DraftTrayId.objects.filter(delink_tray=False)
        if current_batch_id:
            draft_qs = draft_qs.exclude(batch_id__batch_id=current_batch_id)
        draft_trays = list(draft_qs.values_list('tray_id', flat=True).distinct())

        # ── Active submitted trays still sitting in DP (not delinked) ─────────
        # Use DPTrayId_History — trays that moved to IS/Brass QC have
        # DPTrayId_History rows with delink_tray=True after the lot moves on.
        dp_qs = DPTrayId_History.objects.filter(delink_tray=False)
        if current_batch_id:
            dp_qs = dp_qs.exclude(batch_id__batch_id=current_batch_id)
        dp_trays = list(dp_qs.values_list('tray_id', flat=True).distinct())

        all_used_trays = list(set(draft_trays + dp_trays))

        return JsonResponse({
            'success': True,
            'trays': all_used_trays,
            'count': len(all_used_trays)
        })


@method_decorator(csrf_exempt, name='dispatch')
class DraftTrayIdListAPIView(APIView):
    def get(self, request):
        batch_id = request.GET.get('batch_id')
        if not batch_id:
            return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
        
        # ✅ UPDATED: Filter out delinked trays and order by position
        draft_trays = DraftTrayId.objects.filter(
            batch_id__batch_id=batch_id,
            delink_tray=False  # Only get non-delinked trays
        ).order_by('position')
        
        data = [
            {
                'id': tray.id,
                's_no': idx + 1,  # Reindex for display
                'tray_id': tray.tray_id,
                'tray_quantity': tray.tray_quantity,
                'position': idx  # Reset positions for non-delinked trays
            }
            for idx, tray in enumerate(draft_trays)
        ]
        
        return JsonResponse({'success': True, 'trays': data})
    
@method_decorator(csrf_exempt, name='dispatch')
class TrayIdUniqueCheckAPIView(APIView):
    """
    Enhanced TrayId validation - ONLY ALLOW PRE-EXISTING TRAYS from TrayId table
    """
    def get(self, request):
        tray_id = request.GET.get('tray_id')
        batch_id = request.GET.get('batch_id')
        lot_id = request.GET.get('lot_id')
        
        if not batch_id and lot_id:
            return JsonResponse({
                'exists': False,
                'available': False,
                'error': 'Already Scanned',
            }, status=400)
        
        if not tray_id:
            return JsonResponse({'exists': False, 'error': 'Missing tray_id'})

        # Allow partial typing: only error when first two chars are present and not JB/NB
        prefix = tray_id.upper()
        first2 = prefix[:2]
        if len(first2) == 2 and first2 not in ('JB', 'NB'):
            return JsonResponse({
                'exists': False,
                'available': False,
                'error': f'Tray ID "{tray_id}" is not allowed. Only JB and NB trays are permitted.'
            })

        existing_tray = TrayId.objects.filter(tray_id=tray_id).first()

        if not existing_tray:
            return JsonResponse({
                'exists': False,
                'available': False,
                'tray_not_in_system': True,
                'error': f'Tray ID "{tray_id}" not found in system. Only pre-configured trays are allowed.',
                'message': 'This tray must be added by admin before scanning.'
            })

        # ── Cross-check helper: verify tray is genuinely still active in DP ──
        def _is_tray_active_in_dp():
            """Return True if there's a non-delinked DPTrayId_History record for this tray."""
            from .models import DPTrayId_History
            return DPTrayId_History.objects.filter(tray_id=tray_id, delink_tray=False).exists()

        def _release_stale_tray():
            """Reset TrayId master when it's dirty but no active DP record exists."""
            existing_tray.lot_id = None
            existing_tray.batch_id = None
            existing_tray.scanned = False
            existing_tray.delink_tray = False
            existing_tray.save(update_fields=['lot_id', 'batch_id', 'scanned', 'delink_tray'])

        # NEW: Disallow tray if it already has a lot_id value
        if existing_tray.lot_id:
            # Cross-check: if no active DP record exists, the lot_id is stale — reset and allow
            if _is_tray_active_in_dp():
                return JsonResponse({
                    'exists': True,
                    'available': False,
                    'error': f'Tray ID "{tray_id}" is already assigned.',
                    'message': 'This tray is already linked to a lot and cannot be reused until delinked.'
                })
            else:
                _release_stale_tray()
        
        # ✅ NEW: Disallow if tray is rejected
        if getattr(existing_tray, 'rejected_tray', False):
            return JsonResponse({
                'exists': True,
                'available': False,
                'rejected_tray': True,
                'delink_tray': getattr(existing_tray, 'delink_tray', False),  # <-- Add this line
                'error': f'Tray ID \"{tray_id}\" is marked as rejected and cannot be used.',
                'message': 'This tray is rejected and cannot be used for scanning.'
            })
        
        # ✅ NEW: Check if tray is already drafted in another batch (REAL-TIME BLOCKING)
        # This catches the tray immediately while typing, before Submit is clicked
        if batch_id:
            drafted_in_other_batch = DraftTrayId.objects.filter(
                tray_id=tray_id,
                delink_tray=False  # Only block non-delinked drafts
            ).exclude(batch_id__batch_id=batch_id).first()  # Exclude current batch
            
            if drafted_in_other_batch:
                return JsonResponse({
                    'exists': True,
                    'available': False,
                    'already_drafted': True,
                    'error': f'Tray ID "{tray_id}" is already drafted in another batch.',
                    'message': 'This tray has already been reserved in another batch. Cannot reuse until that batch is submitted.'
                })

        # Check if tray is delinked (can be reused regardless of scanned status)
        if existing_tray.delink_tray:
            # Delinked trays can be reused - validate tray type if batch_id provided
            if batch_id:
                validation_result = self.validate_tray_type_compatibility(existing_tray, batch_id)
                if not validation_result['compatible']:
                    return JsonResponse({
                        'exists': True,
                        'available': False,
                        'tray_type_error': True,
                        'delink_tray': True,  # <-- Add this line
                        'error': validation_result['error'],
                        'batch_tray_type': validation_result['batch_tray_type'],
                        'scanned_tray_type': validation_result['scanned_tray_type']
                    })
            
            
            return JsonResponse({
                'exists': True,
                'available': True,
                'delink_tray': True,  # <-- Add this line
                'status': 'delinked_reusable',
                'message': 'Delinked tray - available for reuse'
            })
        
        # Check if tray is already scanned/used
        if existing_tray.scanned:
            # Cross-check: if no active DP record, scanned flag is stale — reset and allow
            if _is_tray_active_in_dp():
                return JsonResponse({
                    'exists': True,
                    'available': False,
                    'already_scanned': True,
                    'delink_tray': getattr(existing_tray, 'delink_tray', False),
                    'error': f'Tray ID "{tray_id}" has already been scanned and is in use',
                    'batch_info': existing_tray.batch_id.batch_id if existing_tray.batch_id else 'Unknown batch',
                    'scan_date': existing_tray.date.strftime('%d-%m-%Y %H:%M') if existing_tray.date else 'Unknown date'
                })
            else:
                _release_stale_tray()
        
        # Tray exists but not scanned - validate tray type compatibility
        if batch_id:
            validation_result = self.validate_tray_type_compatibility(existing_tray, batch_id)
            if not validation_result['compatible']:
                return JsonResponse({
                    'exists': True,
                    'available': False,
                    'tray_type_error': True,
                    'delink_tray': getattr(existing_tray, 'delink_tray', False),  # <-- Add this line
                    'error': validation_result['error'],
                    'batch_tray_type': validation_result['batch_tray_type'],
                    'scanned_tray_type': validation_result['scanned_tray_type']
                })
        
        # Tray is available for use
        return JsonResponse({
            'exists': True,
            'available': True,
            'status': 'pre_configured',
            'message': 'Pre-configured tray - available for scanning',
            'delink_tray': getattr(existing_tray, 'delink_tray', False),  # <-- Add this line
            'tray_type': existing_tray.tray_type,
            'tray_capacity': existing_tray.tray_capacity
        })
    
    def validate_tray_type_compatibility(self, tray, batch_id):
        """
        Validate if the scanned tray type is compatible with the batch tray type
        
        Args:
            tray: TrayId object with tray_type and tray_capacity
            batch_id: Batch ID string
            
        Returns:
            dict: {
                'compatible': bool,
                'error': str,
                'batch_tray_type': str,
                'scanned_tray_type': str
            }
        """
        try:
            # Get batch instance
            batch_instance = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not batch_instance:
                return {
                    'compatible': False,
                    'error': 'Batch not found',
                    'batch_tray_type': None,
                    'scanned_tray_type': tray.tray_type
                }
            
            # Get batch tray type
            batch_tray_type = batch_instance.tray_type
            scanned_tray_type = tray.tray_type
            
            print(f"🔍 Tray Type Validation: Batch={batch_tray_type}, Scanned Tray={scanned_tray_type}")
            
            # If either tray type is not set, allow but warn
            if not batch_tray_type or not scanned_tray_type:
                return {
                    'compatible': True,  # Allow if not configured
                    'error': None,
                    'batch_tray_type': batch_tray_type,
                    'scanned_tray_type': scanned_tray_type
                }
            
            # Normalize both to pre-jig category (normal/jumbo) for comparison
            def _norm_cat(tt_str):
                tt = (tt_str or '').upper()
                return 'jumbo' if ('JUMBO' in tt or tt in ('JR', 'JD', 'JB', 'JL')) else 'normal'

            batch_category = _norm_cat(batch_tray_type)
            scanned_category = _norm_cat(scanned_tray_type)

            if batch_category != scanned_category:
                error_msg = f"❌ Tray Type Mismatch: Batch requires '{batch_category}' type but scanned tray '{tray.tray_id}' is '{scanned_category}' type"
                return {
                    'compatible': False,
                    'error': error_msg,
                    'batch_tray_type': batch_tray_type,
                    'scanned_tray_type': scanned_tray_type
                }

            # Validate pre-jig capacity: Jumbo=12, Normal=16
            expected_prejig_cap = 12 if batch_category == 'jumbo' else 16
            scanned_tray_capacity = tray.tray_capacity

            if scanned_tray_capacity and scanned_tray_capacity != expected_prejig_cap:
                error_msg = f"⚠️ Tray Capacity Mismatch: Expected {expected_prejig_cap} for {batch_category} type, but scanned tray has capacity {scanned_tray_capacity}"
                return {
                    'compatible': False,
                    'error': error_msg,
                    'batch_tray_type': batch_tray_type,
                    'scanned_tray_type': scanned_tray_type
                }

            # All validations passed
            return {
                'compatible': True,
                'error': None,
                'batch_tray_type': batch_tray_type,
                'scanned_tray_type': scanned_tray_type
            }
            
        except Exception as e:
            logger.error(f"❌ Error in tray type validation: {str(e)}", exc_info=True)
            return {
                'compatible': False,
                'error': 'Unable to process the request. Please verify the submitted data and try again.',
                'batch_tray_type': None,
                'scanned_tray_type': tray.tray_type if tray else None
            }



# Add this new API view to your views.py file

@method_decorator(csrf_exempt, name='dispatch')
class UpdateBatchQuantityAndColorAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            new_quantity = data.get('total_batch_quantity')
            new_plating_color = data.get('plating_color')
            
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            
            obj = ModelMasterCreation.objects.filter(batch_id=batch_id, Moved_to_D_Picker=False).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found or already moved'}, status=404)
            
            # Update quantity if provided
            if new_quantity is not None:
                try:
                    new_quantity = int(new_quantity)
                    if new_quantity <= 0:
                        return JsonResponse({'success': False, 'error': 'Quantity must be greater than 0'}, status=400)
                    
                    # ✅ FIXED: If quantity changes, invalidate top tray verification
                    if obj.total_batch_quantity != new_quantity:
                        obj.top_tray_qty_verified = False
                        print(f"🔄 Quantity changed from {obj.total_batch_quantity} to {new_quantity} - resetting top_tray_qty_verified to False")
                    
                    obj.total_batch_quantity = new_quantity
                except (ValueError, TypeError):
                    return JsonResponse({'success': False, 'error': 'Invalid quantity value'}, status=400)
            
            # Update plating color if provided
            if new_plating_color:
                # Validate that the plating color exists in Plating_Color table
                plating_color_obj = Plating_Color.objects.filter(plating_color=new_plating_color).first()
                if not plating_color_obj:
                    available_colors = list(Plating_Color.objects.values_list('plating_color', flat=True)[:5])
                    return JsonResponse({
                        'success': False, 
                        'error': f'Plating color "{new_plating_color}" not found in Master Data',
                        'available_colors': available_colors
                    }, status=400)
                obj.plating_color = plating_color_obj
            
            # Save the changes
            update_fields = []
            if new_quantity is not None:
                update_fields.extend(['total_batch_quantity', 'top_tray_qty_verified'])
            if new_plating_color:
                update_fields.append('plating_color')
            
            if update_fields:
                obj.save(update_fields=update_fields)
            
            return JsonResponse({
                'success': True, 
                'message': 'Record updated successfully',
                'updated_quantity': obj.total_batch_quantity,
                'updated_plating_color': obj.plating_color
            })
            
        except Exception as e:
            logger.error(f"❌ Error in UpdateBatchQuantityAndColorAPIView: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class GetPlatingColorsAPIView(APIView):
    """
    API endpoint to fetch all plating colors for dropdown navigation
    """
    def get(self, request):
        try:
            plating_colors = list(Plating_Color.objects.values_list('plating_color', flat=True).order_by('plating_color'))
            return JsonResponse({
                'success': True,
                'plating_colors': plating_colors,
                'count': len(plating_colors)
            })
        except Exception as e:
            logger.error(f"❌ Error in GetPlatingColorsAPIView: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)



@method_decorator(csrf_exempt, name='dispatch')
class DeleteBatchAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            obj = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found'}, status=404)
            obj.delete()
            return JsonResponse({'success': True, 'message': 'Batch deleted'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class SaveDPPickRemarkAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            remark = data.get('remark', '').strip()
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            obj = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
            if not obj:
                return JsonResponse({'success': False, 'error': 'Batch not found'}, status=404)
            obj.dp_pick_remarks = remark
            obj.save(update_fields=['dp_pick_remarks'])
            return JsonResponse({'success': True, 'message': 'Remark saved'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)

@method_decorator(csrf_exempt, name='dispatch')
class TrayValidateAPIView(APIView):
    def post(self, request):
        try:
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id_input = str(data.get('batch_id')).strip()
            tray_id = str(data.get('tray_id')).strip()
            print(f"[TrayValidateAPIView] User entered: batch_id={batch_id_input}, tray_id={tray_id}")

            # Try to match batch_id by code part (after last '-')
            trays = DPTrayId_History.objects.filter(batch_id__batch_id__icontains=batch_id_input)
            print(f"A   ll tray_ids for batch containing '{batch_id_input}': {[t.tray_id for t in trays]}")

            exists = trays.filter(tray_id=tray_id).exists()
            print(f"Exists in TrayId table? {exists}")

            return JsonResponse({'success': True, 'exists': exists})
        except Exception as e:
            logger.error(f"[TrayValidateAPIView] Error: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)
     
     
import pytz

class DPCompletedTableView(APIView):
    renderer_classes = [TemplateHTMLRenderer]
    template_name = 'Day_Planning/DP_Completed_Table.html'
    permission_classes = [IsAuthenticated] 

    def get(self, request, *args, **kwargs):
        from django.utils import timezone

        user = request.user
        tz = pytz.timezone("Asia/Kolkata")
        now_local = timezone.now().astimezone(tz)
        today = now_local.date()
        yesterday = today - timedelta(days=1)

        # --- Use created_at for date filtering ---
        # Get all related created_at values for completed batches
        completed_batches = ModelMasterCreation.objects.filter(
            total_batch_quantity__gt=0,
            Moved_to_D_Picker=True
        ).values_list('batch_id', flat=True)

        # Get min/max created_at from TotalStockModel for these batches
        created_at_qs = TotalStockModel.objects.filter(
            batch_id__batch_id__in=completed_batches
        )
        min_created_at = created_at_qs.order_by('created_at').values_list('created_at', flat=True).first()
        max_created_at = created_at_qs.order_by('-created_at').values_list('created_at', flat=True).first()

        # Always use current date in IST
        today = now_local.date()
        yesterday = today - timedelta(days=1) 

        # Get date filter parameters from request
        from_date_str = request.GET.get('from_date')
        to_date_str = request.GET.get('to_date')

        default_from_date = min_created_at.astimezone(tz).date() if min_created_at else yesterday
        default_to_date = max_created_at.astimezone(tz).date() if max_created_at else today
        date_filter_applied = bool(from_date_str and to_date_str)

        # Calculate date range. Without an explicit search, show all completed lots.
        if date_filter_applied:
            try:
                from_date = _dt.datetime.strptime(from_date_str, '%Y-%m-%d').date()
                to_date = _dt.datetime.strptime(to_date_str, '%Y-%m-%d').date()
            except ValueError:
                from_date = default_from_date
                to_date = default_to_date
                date_filter_applied = False
        else:
            from_date = default_from_date
            to_date = default_to_date

        # Convert dates to datetime objects for filtering (include full day)
        from_datetime = timezone.make_aware(_dt.datetime.combine(from_date, _dt.datetime.min.time()))
        to_datetime = timezone.make_aware(_dt.datetime.combine(to_date, _dt.datetime.max.time()))
        # Subqueries for annotations
        last_process_module_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('last_process_module')[:1]
        next_process_module_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('next_process_module')[:1]
        created_at_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk')
        ).values('created_at')[:1]
        accepted_Ip_stock_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk'),
        ).values('accepted_Ip_stock')[:1]
        few_cases_accepted_Ip_stock_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk'),
        ).values('few_cases_accepted_Ip_stock')[:1]
        rejected_ip_stock_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk'),
        ).values('rejected_ip_stock')[:1]
        ip_person_qty_verified_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk'),
        ).values('ip_person_qty_verified')[:1]
        draft_tray_verify_subquery = TotalStockModel.objects.filter(
            batch_id=OuterRef('pk'),
        ).values('draft_tray_verify')[:1]

        queryset = ModelMasterCreation.objects.filter(
            total_batch_quantity__gt=0,
            Moved_to_D_Picker=True,
        ).annotate(
            last_process_module=Subquery(last_process_module_subquery),
            next_process_module=Subquery(next_process_module_subquery),
            created_at=Subquery(created_at_subquery),
            accepted_Ip_stock=Subquery(accepted_Ip_stock_subquery),
            ip_person_qty_verified=Subquery(ip_person_qty_verified_subquery),
            few_cases_accepted_Ip_stock=Subquery(few_cases_accepted_Ip_stock_subquery),
            rejected_ip_stock=Subquery(rejected_ip_stock_subquery),
            draft_tray_verify=Subquery(draft_tray_verify_subquery),
        )
        # Exclude batches that have no TotalStockModel record (created_at is null)
        # This avoids showing deleted/partial records which were removed downstream
        queryset = queryset.filter(created_at__isnull=False)

        if date_filter_applied:
            batch_ids_in_range = TotalStockModel.objects.filter(
                created_at__range=(from_datetime, to_datetime)
            ).values_list('batch_id__batch_id', flat=True)
            queryset = queryset.filter(batch_id__in=batch_ids_in_range)

        queryset = queryset.order_by('-created_at')

        # Pagination
        page_number = request.GET.get('page', 1)
        paginator = Paginator(queryset, 10)
        page_obj = paginator.get_page(page_number)

        master_data = list(page_obj.object_list.values(
            'batch_id',
            'date_time',
            'model_stock_no__model_no',
            'plating_color',
            'polish_finish',
            'version__version_name',
            'vendor_internal',
            'location__location_name',
            'no_of_trays',
            'tray_type',
            'total_batch_quantity',
            'tray_capacity',
            'Moved_to_D_Picker',
            'last_process_module',
            'next_process_module',
            'Draft_Saved',
            'top_tray_qty_verified',
            'created_at',
            'plating_stk_no',
            'polishing_stk_no',
            'category',
            'version__version_internal',
            'dp_pick_remarks',
            'accepted_Ip_stock',
            'rejected_ip_stock',
            'few_cases_accepted_Ip_stock',
            'ip_person_qty_verified',
            'draft_tray_verify'
        ))

        # ✅ PERF: Batch-fetch model images ONCE for the whole page instead of a
        # per-row ModelMasterCreation query inside the loop (was N+1).
        page_batch_ids = [d['batch_id'] for d in master_data]
        images_by_batch = {}
        for mmc in ModelMasterCreation.objects.filter(
            batch_id__in=page_batch_ids
        ).prefetch_related('images'):
            urls = [
                img.master_image.url
                for img in mmc.images.all()
                if getattr(img, 'master_image', None)
            ]
            images_by_batch[mmc.batch_id] = urls

        # Calculate no_of_trays dynamically and add tray_qty_list
        for data in master_data:
            total_batch_quantity = data.get('total_batch_quantity', 0)
            tray_capacity = data.get('tray_capacity', 0)
            data['vendor_location'] = f"{data.get('vendor_internal', '')}_{data.get('location__location_name', '')}"
            
            if tray_capacity > 0:
                no_of_trays = math.ceil(total_batch_quantity / tray_capacity)
                data['no_of_trays'] = no_of_trays
                
                # Calculate tray_qty_list (same logic as DP_PickTable)
                tray_qty_list = []
                remainder = total_batch_quantity % tray_capacity
                if no_of_trays == 1:
                    tray_qty_list = [total_batch_quantity]
                elif no_of_trays > 1:
                    if remainder != 0:
                        tray_qty_list.append(remainder)
                        for _ in range(1, no_of_trays):
                            tray_qty_list.append(tray_capacity)
                    else:
                        for _ in range(no_of_trays):
                            tray_qty_list.append(tray_capacity)
                data['tray_qty_list'] = tray_qty_list
            else:
                data['no_of_trays'] = 0
                data['tray_qty_list'] = []

            # Add model images (precomputed once above — no per-row query).
            images = list(images_by_batch.get(data['batch_id'], []))
            if not images:
                from django.templatetags.static import static
                images = [static('assets/images/imagePlaceholder.jpg')]
            data['model_images'] = images

        context = {
            'master_data': master_data,
            'page_obj': page_obj,
            'paginator': paginator,
            'user': user,
            'from_date': from_date.strftime('%Y-%m-%d'),  # 🔥 NEW: Pass dates to template
            'to_date': to_date.strftime('%Y-%m-%d'),      # 🔥 NEW: Pass dates to template
            'date_filter_applied': date_filter_applied,
        }
        return Response(context, template_name=self.template_name)
    
    
    
@method_decorator(csrf_exempt, name='dispatch')
class CompletedTrayIdListAPIView(APIView):
    def get(self, request):
        batch_id = request.GET.get('batch_id')
        if not batch_id:
            return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
        
        # ✅ FIX: Fetch lot quantity from ModelMasterCreation
        batch_obj = ModelMasterCreation.objects.filter(batch_id=batch_id).first()
        lot_quantity = batch_obj.total_batch_quantity if batch_obj else 0
        
        top_tray = DPTrayId_History.objects.filter(
            batch_id__batch_id=batch_id,
            top_tray=True,
            rejected_tray=False,
            tray_quantity__gt=0
        ).first()
        
        other_trays = DPTrayId_History.objects.filter(
            batch_id__batch_id=batch_id,
            top_tray=False,
            rejected_tray=False,
            tray_quantity__gt=0
        ).order_by('id')
        
        data = []
        if top_tray:
            data.append({
                's_no': 1,
                'tray_id': top_tray.tray_id,
                'tray_quantity': top_tray.tray_quantity,
                'position': 0,
                'is_top_tray': True,
                'delink_tray': top_tray.delink_tray,  # <-- ADD THIS
            })
        for idx, tray in enumerate(other_trays):
            data.append({
                's_no': idx + 2,
                'tray_id': tray.tray_id,
                'tray_quantity': tray.tray_quantity,
                'position': idx + 1,
                'is_top_tray': False,
                'delink_tray': tray.delink_tray,  # <-- ADD THIS
            })
        
        # ✅ FIX: Include lot quantity in response
        return JsonResponse({'success': True, 'trays': data, 'lot_quantity': lot_quantity})
   
@method_decorator(csrf_exempt, name='dispatch')
class TrayAutoSaveAPIView(APIView):
    """
    API for cross-browser tray auto-save functionality
    """
    
    def post(self, request):
        """Save auto-save data for current user and batch"""
        try:
            if not request.user.is_authenticated:
                return JsonResponse({'success': False, 'error': 'User not authenticated'}, status=401)
            
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            batch_id = data.get('batch_id')
            tray_data = data.get('tray_data', [])
            modal_data = data.get('modal_data', {})
            
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            
            # Check if there's valid data to save
            has_valid_data = any(
                tray.get('trayId', '').strip() or tray.get('trayQty', '')
                for tray in tray_data
            )
            
            if not has_valid_data:
                # No valid data - delete any existing auto-save
                TrayAutoSaveData.objects.filter(user=request.user, batch_id=batch_id).delete()
                return JsonResponse({'success': True, 'message': 'Auto-save cleared (no data)'})
            
            # Prepare auto-save data
            auto_save_data = {
                'timestamp': timezone.now().isoformat(),
                'batch_id': batch_id,
                'tray_data': tray_data,
                'modal_data': modal_data,
                'user_agent': request.META.get('HTTP_USER_AGENT', '')[:200],  # Track browser
            }
            
            # Update or create auto-save record
            auto_save_obj, created = TrayAutoSaveData.objects.update_or_create(
                user=request.user,
                batch_id=batch_id,
                defaults={'auto_save_data': auto_save_data}
            )
            
            action = 'created' if created else 'updated'
            print(f"✅ Auto-save {action} for user {request.user.username}, batch {batch_id}")
            
            return JsonResponse({
                'success': True, 
                'message': f'Auto-save {action}',
                'data_id': auto_save_obj.id
            })
            
        except Exception as e:
            logger.error(f"❌ Error in TrayAutoSaveAPIView POST: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)
    
    def get(self, request):
        """Retrieve auto-save data for current user and batch"""
        try:
            if not request.user.is_authenticated:
                return JsonResponse({'success': False, 'error': 'User not authenticated'}, status=401)
            
            batch_id = request.GET.get('batch_id')
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            
            # Get auto-save data for this user and batch
            auto_save_obj = TrayAutoSaveData.objects.filter(
                user=request.user, 
                batch_id=batch_id
            ).first()
            
            if not auto_save_obj:
                return JsonResponse({
                    'success': True, 
                    'has_data': False, 
                    'message': 'No auto-save data found'
                })
            
            # Check if data is expired (older than 24 hours)
            if auto_save_obj.is_expired(hours=24):
                auto_save_obj.delete()
                return JsonResponse({
                    'success': True, 
                    'has_data': False, 
                    'message': 'Auto-save data expired and cleared'
                })
            
            # Return the auto-save data
            return JsonResponse({
                'success': True,
                'has_data': True,
                'data': auto_save_obj.auto_save_data,
                'saved_at': auto_save_obj.updated_at.isoformat(),
                'user': request.user.username
            })
            
        except Exception as e:
            logger.error(f"❌ Error in TrayAutoSaveAPIView GET: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)
    
    def delete(self, request):
        """Clear auto-save data for current user and batch"""
        try:
            if not request.user.is_authenticated:
                return JsonResponse({'success': False, 'error': 'User not authenticated'}, status=401)
            
            batch_id = request.GET.get('batch_id')
            if not batch_id:
                return JsonResponse({'success': False, 'error': 'Missing batch_id'}, status=400)
            
            # Delete auto-save data
            deleted_count, _ = TrayAutoSaveData.objects.filter(
                user=request.user, 
                batch_id=batch_id
            ).delete()
            
            return JsonResponse({
                'success': True, 
                'message': f'Auto-save cleared ({deleted_count} records deleted)'
            })
            
        except Exception as e:
            logger.error(f"❌ Error in TrayAutoSaveAPIView DELETE: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)



@method_decorator(csrf_exempt, name='dispatch')
class TrayAutoSaveCleanupAPIView(APIView):
    """
    API to clean up old auto-save data (optional - can be called periodically)
    """
    
    def post(self, request):
        """Clean up auto-save data older than specified hours"""
        try:
            if not request.user.is_authenticated or not request.user.is_staff:
                return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
            
            data = request.data if hasattr(request, 'data') else json.loads(request.body.decode('utf-8'))
            hours = data.get('hours', 72)  # Default: 72 hours (3 days)
            
            # Delete old auto-save data
            cutoff_time = timezone.now() - timedelta(hours=hours)
            
            deleted_count, _ = TrayAutoSaveData.objects.filter(
                updated_at__lt=cutoff_time
            ).delete()
            
            return JsonResponse({
                'success': True, 
                'message': f'Cleaned up {deleted_count} old auto-save records (older than {hours} hours)'
            })
            
        except Exception as e:
            logger.error(f"❌ Error in TrayAutoSaveCleanupAPIView: {str(e)}", exc_info=True)
            return JsonResponse({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'}, status=500)
            
# ==========================================
# BARCODE SCANNER API - Day Planning
# ==========================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dget_lot_id_for_tray(request):
    """
    API to get lot_id for a scanned tray_id in Day Planning module
    Enhanced with multiple lookup sources and better debugging
    """
    tray_id = request.GET.get('tray_id', '').strip()
    if not tray_id:
        return JsonResponse({'success': False, 'error': 'Tray ID required'})
    
    try:
        print(f"🔍 [DP SCANNER] Looking up tray_id: {tray_id}")
        
        # Helper function to calculate page number for a batch
        def calculate_page_number(batch_id):
            try:
                # Build the same queryset as in DayPlanningPickTableAPIView
                # Subqueries for annotations
                last_process_module_subquery = TotalStockModel.objects.filter(
                    batch_id=OuterRef('pk')
                ).values('last_process_module')[:1]
                next_process_module_subquery = TotalStockModel.objects.filter(
                    batch_id=OuterRef('pk')
                ).values('next_process_module')[:1]
                accepted_Ip_stock_subquery = TotalStockModel.objects.filter(
                    batch_id=OuterRef('pk'),
                ).values('accepted_Ip_stock')[:1]   
                tray_scan_status_subquery = TotalStockModel.objects.filter(
                    batch_id=OuterRef('pk')
                ).values('tray_scan_status')[:1]
                
                # Build exact same queryset as the view
                queryset = ModelMasterCreation.objects.filter(
                    total_batch_quantity__gt=0,
                    Moved_to_D_Picker=False  
                ).annotate(
                    last_process_module=Subquery(last_process_module_subquery),
                    next_process_module=Subquery(next_process_module_subquery),
                    accepted_Ip_stock=Subquery(accepted_Ip_stock_subquery),
                    tray_scan_status=Subquery(tray_scan_status_subquery),
                ).order_by('-date_time', 'batch_id')  # Default sorting from view
                
                # Find the position of the target batch in the ordered queryset
                batch_ids = list(queryset.values_list('batch_id', flat=True))
                print(f"🔍 [DP SCANNER] Total batches in pick table: {len(batch_ids)}")
                
                try:
                    batch_position = batch_ids.index(batch_id) + 1  # 1-based position
                    page_number = ((batch_position - 1) // 10) + 1  # 10 items per page
                    print(f"🔍 [DP SCANNER] Batch {batch_id} is at position {batch_position}, page {page_number}")
                    return page_number
                except ValueError:
                    print(f"⚠️ [DP SCANNER] Batch {batch_id} not found in pick table queryset")
                    return None
                    
            except Exception as e:
                logger.error(f"❌ [DP SCANNER] Error calculating page: {str(e)}", exc_info=True)
                return None
        
        # 1. Look up in DPTrayId_History table (Day Planning specific)
        dp_tray = DPTrayId_History.objects.filter(tray_id=tray_id).first()
        if dp_tray and dp_tray.lot_id:
            print(f"✅ [DP SCANNER] Found in DPTrayId_History: lot_id={dp_tray.lot_id}")
            batch_tray_capacity = dp_tray.batch_id.tray_capacity if dp_tray.batch_id else None
            page_number = calculate_page_number(dp_tray.batch_id.batch_id) if dp_tray.batch_id else None
            return JsonResponse({
                'success': True,
                'lot_id': dp_tray.lot_id,
                'batch_id': dp_tray.batch_id.batch_id if dp_tray.batch_id else None,
                'tray_quantity': dp_tray.tray_quantity,
                'tray_capacity': dp_tray.tray_capacity,
                'batch_tray_capacity': batch_tray_capacity,
                'page': page_number,
                'context': 'dp_tray_history'
            })
        
        # 2. Look up in TrayId table (main tray storage)
        tray_obj = TrayId.objects.filter(tray_id=tray_id).first()
        
        if tray_obj and tray_obj.batch_id:
            print(f"✅ [DP SCANNER] Found in TrayId: batch_id={tray_obj.batch_id.batch_id}")
            page_number = calculate_page_number(tray_obj.batch_id.batch_id) if tray_obj.batch_id else None
            return JsonResponse({
                'success': True,
                'lot_id': tray_obj.lot_id,
                'batch_id': tray_obj.batch_id.batch_id if tray_obj.batch_id else None,
                'tray_quantity': tray_obj.tray_quantity,
                'tray_capacity': tray_obj.tray_capacity,
                'batch_tray_capacity': tray_obj.batch_id.tray_capacity if tray_obj.batch_id else None,
                'page': page_number,
                'context': 'tray_id_table'
            })
        
        # 3. Look up in DraftTrayId table (for draft entries)
        draft_tray = DraftTrayId.objects.filter(tray_id=tray_id).first()
        
        if draft_tray and draft_tray.batch_id:
            print(f"✅ [DP SCANNER] Found in DraftTrayId: batch_id={draft_tray.batch_id.batch_id}")
            page_number = calculate_page_number(draft_tray.batch_id.batch_id) if draft_tray.batch_id else None
            return JsonResponse({
                'success': True,
                'lot_id': draft_tray.lot_id,
                'batch_id': draft_tray.batch_id.batch_id if draft_tray.batch_id else None,
                'tray_quantity': draft_tray.tray_quantity,
                'tray_capacity': None,  # DraftTrayId doesn't have capacity
                'batch_tray_capacity': draft_tray.batch_id.tray_capacity if draft_tray.batch_id else None,
                'page': page_number,
                'context': 'draft_tray_table'
            })
        
        # 4. Check if tray exists but has no batch_id assigned
        if dp_tray or tray_obj or draft_tray:
            print(f"⚠️ [DP SCANNER] Tray exists but no batch_id assigned")
            return JsonResponse({
                'success': False, 
                'error': 'Tray exists but not assigned to any batch'
            })
        
        # 5. Tray not found anywhere
        print(f"❌ [DP SCANNER] Tray not found in any table")
        return JsonResponse({
            'success': False, 
            'error': f'Tray ID {tray_id} not found in system'
        })
        
    except Exception as e:
        logger.error(f"❌ [DP SCANNER] Error: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False, 
            'error': 'Unable to process the request. Please verify the submitted data and try again.'
        })


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def draft_tray_delete(request):
    """
    API to delete a draft tray when input is cleared
    """
    batch_id = request.GET.get('batch_id')
    position = request.GET.get('position')
    if not batch_id or not position:
        return JsonResponse({'success': False, 'error': 'Batch ID and position required'})
    
    try:
        # Get the batch object
        batch = ModelMasterCreation.objects.get(batch_id=batch_id)
        # Delete the draft tray
        deleted = DraftTrayId.objects.filter(batch_id=batch, position=position).delete()
        print(f"🗑️ [DRAFT DELETE] Deleted {deleted[0]} draft tray(s) for batch {batch_id} position {position}")
        return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f"❌ [DRAFT DELETE] Error: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Unable to process the request. Please verify the submitted data and try again.'})


class GetAllowedVersionsAPIView(APIView):
    """
    API endpoint to fetch all allowed version codes from the database.
    This provides dynamic validation instead of hardcoded frontend arrays.
    """
    
    def get(self, request):
        try:
            # Get all unique version codes from the database
            version_codes = set()
            
            # Get all versions and extract both version_name and version_internal
            versions = Version.objects.all()
            for version in versions:
                if version.version_name:
                    version_codes.add(version.version_name.strip())
                if version.version_internal:
                    version_codes.add(version.version_internal.strip())
            
            # Convert to sorted list for consistent output
            allowed_versions = sorted(list(version_codes))
            
            return JsonResponse({
                'success': True,
                'allowed_versions': allowed_versions
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': 'Unable to process the request. Please verify the submitted data and try again.',
                'allowed_versions': ['A', 'B', 'C', 'D', 'E', 'L']  # Fallback
            })


class ValidatePlatingStockNoAPIView(APIView):
    """
    API endpoint to validate plating stock number against ModelMaster database.
    Checks if the plating stock number exists in the database.
    """
    # Simple in-memory cache for plating stock numbers
    _cache = {}
    
    def get(self, request):
        plating_stk_no = request.GET.get('plating_stk_no', '').strip()
        
        if not plating_stk_no:
            return JsonResponse({
                'success': False,
                'is_valid': False,
                'message': 'Plating stock number is required'
            })
        
        # Check cache first
        if plating_stk_no in self._cache:
            cached_result = self._cache[plating_stk_no]
            return JsonResponse(cached_result)
        
        try:
            # Check if plating stock number exists in ModelMaster
            model = ModelMaster.objects.get(plating_stk_no=plating_stk_no)
            result = {
                'success': True,
                'is_valid': True,
                'message': 'Plating stock number found',
                'data': {
                    'model_no': model.model_no,
                    'version': model.version,
                    'brand': model.brand,
                    'ep_bath_type': model.ep_bath_type,
                }
            }
            # Cache the result
            self._cache[plating_stk_no] = result
            return JsonResponse(result)
        except ModelMaster.DoesNotExist:
            result = {
                'success': True,
                'is_valid': False,
                'message': f'Plating stock number "{plating_stk_no}" not found in database. Please check the number or add it via Django Admin.'
            }
            # Cache the result
            self._cache[plating_stk_no] = result
            return JsonResponse(result)
                
        except Exception as e:
            result = {
                'success': False,
                'is_valid': False,
                'message': 'Unable to process the request. Please verify the submitted data and try again.'
            }
            return JsonResponse(result)

from django.http import HttpResponse
from io import BytesIO
from datetime import datetime

class DownloadExcelTemplateAPIView(APIView):
    """
    API endpoint to download Excel template for Day Planning bulk upload
    
    Template structure:
    - Headers: S.No, Plating Stk No, Polishing Stk No, Plating Colour, Category, Input Qty, Source
    - Sample row with example data
    - Formatted with headers styling
    - No data rows - user fills their own data
    
    Performance:
    - Uses BytesIO for in-memory buffer (no disk I/O)
    - Single-pass row creation
    - Cached in response headers for browser
    """
    
    def get(self, request):
        """Generate and return Excel template"""
        try:
            # Create a new workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bulk Upload"
            
            # Define columns (matching the data table structure)
            columns = ['S.No', 'Plating Stk No', 'Polishing Stk No', 'Plating Colour', 'Category', 'Input Qty', 'Source']
            
            # Add header row
            for col_num, column_title in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = column_title
                # Style header: bold, light blue background, centered
                from openpyxl.styles import Font, PatternFill, Alignment
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="126A83", end_color="126A83", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths for better readability
            ws.column_dimensions['A'].width = 8     # S.No
            ws.column_dimensions['B'].width = 15    # Plating Stk No
            ws.column_dimensions['C'].width = 15    # Polishing Stk No
            ws.column_dimensions['D'].width = 15    # Plating Colour
            ws.column_dimensions['E'].width = 15    # Category
            ws.column_dimensions['F'].width = 12    # Input Qty
            ws.column_dimensions['G'].width = 15    # Source
            
            # Freeze the header row
            ws.freeze_panes = "A2"
            
            # Create BytesIO object to hold Excel file in memory
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Create HTTP response with proper headers
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # File name with timestamp
            timestamp = datetime.now().strftime('%d%b%Y')
            response['Content-Disposition'] = f'attachment; filename="DayPlanning_Bulk_Upload_Template_{timestamp}.xlsx"'
            
            return response
            
        except Exception as e:
            logger.error(f"❌ Error generating Excel template: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': 'Unable to process the request. Please verify the submitted data and try again.'
            }, status=500)